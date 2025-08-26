
import re
import json
import csv
import io
from typing import List, Dict, Any, Optional, TypedDict, Annotated, Union
from dataclasses import dataclass, asdict
from enum import Enum
import operator
from datetime import datetime

# LangGraph imports
from langgraph.graph import StateGraph, END
from langchain_core.messages import BaseMessage, HumanMessage, AIMessage
from langchain_core.prompts import ChatPromptTemplate
from langchain_openai import ChatOpenAI

# Environment variable loading
import os
from dotenv import load_dotenv

# Load environment variables from .env file
load_dotenv()

# Access the key
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")
if not OPENAI_API_KEY:
    raise ValueError("OPENAI_API_KEY is not set in environment variables")

class TestType(Enum):
    POSITIVE = "positive"
    NEGATIVE = "negative"
    BOUNDARY = "boundary"
    EDGE_CASE = "edge_case"
    INTEGRATION = "integration"
    SECURITY = "security"
    PERFORMANCE = "performance"
    USABILITY = "usability"

class RequirementSource(Enum):
    MANUAL = "manual"
    CSV_FILE = "csv_file"
    EXCEL_FILE = "excel_file"
    JSON_FILE = "json_file"
    TEXT_FILE = "text_file"
    WORD_FILE = "word_file"

class RequirementType(Enum):
    FUNCTIONAL = "functional"
    NON_FUNCTIONAL = "non_functional"
    SECURITY = "security"
    PERFORMANCE = "performance"
    USABILITY = "usability"
    INTEGRATION = "integration"

@dataclass
class Requirement:
    """Represents a single requirement"""
    id: str
    text: str
    priority: str = "Medium"
    category: str = "Functional"
    source: RequirementSource = RequirementSource.MANUAL
    line_number: Optional[int] = None
    
@dataclass
class TestCase:
    """Represents a single test case"""
    id: str
    title: str
    description: str
    test_type: TestType
    preconditions: List[str]
    test_steps: List[str]
    expected_result: str
    priority: str = "Medium"
    tags: List[str] = None
    risk_level: str = "Medium"
    automation_feasible: bool = True
    requirement_id: str = ""
    estimated_time: int = 30
    test_data: List[str] = None
    
    def __post_init__(self):
        if self.tags is None:
            self.tags = []
        if self.test_data is None:
            self.test_data = []

class FileProcessor:
    """Handles processing of various file formats containing requirements"""
    
    @staticmethod
    def process_file(file_path: str) -> List[Requirement]:
        """Process a file and extract requirements based on file type"""
        try:
            file_extension = file_path.lower().split('.')[-1]
            
            if file_extension == 'csv':
                return FileProcessor._process_csv_file(file_path)
            elif file_extension in ['xlsx', 'xls']:
                return FileProcessor._process_excel_file(file_path)
            elif file_extension == 'json':
                return FileProcessor._process_json_file(file_path)
            elif file_extension == 'txt':
                return FileProcessor._process_text_file(file_path)
            elif file_extension in ['docx', 'doc']:
                return FileProcessor._process_word_file(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
                
        except Exception as e:
            print(f"Error processing file {file_path}: {e}")
            return []
    
    @staticmethod
    def _process_csv_file(file_path: str) -> List[Requirement]:
        """Process CSV file containing requirements"""
        requirements = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
        except Exception as e:
            print(f"Error reading CSV file: {e}")
            return []
        
        # Parse CSV content
        csv_reader = csv.DictReader(io.StringIO(content))
        
        for row_num, row in enumerate(csv_reader, 1):
            # Try different possible column names
            req_id = (row.get('ID') or row.get('id') or 
                     row.get('Requirement ID') or row.get('req_id') or 
                     f"REQ{row_num:03d}")
            
            req_text = (row.get('Requirement') or row.get('requirement') or 
                       row.get('Description') or row.get('description') or 
                       row.get('Text') or row.get('text') or "")
            
            priority = (row.get('Priority') or row.get('priority') or "Medium")
            category = (row.get('Category') or row.get('category') or "Functional")
            
            if req_text.strip():
                requirement = Requirement(
                    id=req_id,
                    text=req_text.strip(),
                    priority=priority,
                    category=category,
                    source=RequirementSource.CSV_FILE,
                    line_number=row_num
                )
                requirements.append(requirement)
        
        return requirements
    
    @staticmethod
    def _process_excel_file(file_path: str) -> List[Requirement]:
        """Process Excel file containing requirements"""
        requirements = []
        
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            # Find header row
            headers = {}
            for col in range(1, worksheet.max_column + 1):
                header_value = worksheet.cell(1, col).value
                if header_value:
                    header_lower = str(header_value).lower()
                    if 'id' in header_lower:
                        headers['id'] = col
                    elif any(x in header_lower for x in ['requirement', 'description', 'text']):
                        headers['text'] = col
                    elif 'priority' in header_lower:
                        headers['priority'] = col
                    elif 'category' in header_lower:
                        headers['category'] = col
            
            # Process data rows
            for row_num in range(2, worksheet.max_row + 1):
                req_id = str(worksheet.cell(row_num, headers.get('id', 1)).value or f"REQ{row_num-1:03d}")
                req_text = str(worksheet.cell(row_num, headers.get('text', 2)).value or "")
                priority = str(worksheet.cell(row_num, headers.get('priority', 3)).value or "Medium")
                category = str(worksheet.cell(row_num, headers.get('category', 4)).value or "Functional")
                
                if req_text.strip() and req_text != "None":
                    requirement = Requirement(
                        id=req_id,
                        text=req_text.strip(),
                        priority=priority,
                        category=category,
                        source=RequirementSource.EXCEL_FILE,
                        line_number=row_num
                    )
                    requirements.append(requirement)
                    
        except ImportError:
            print("openpyxl not installed. Install with: pip install openpyxl")
        except Exception as e:
            print(f"Error processing Excel file: {e}")
        
        return requirements
    
    @staticmethod
    def _process_json_file(file_path: str) -> List[Requirement]:
        """Process JSON file containing requirements"""
        requirements = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            
            # Handle different JSON structures
            if isinstance(data, list):
                # Array of requirement objects
                for i, item in enumerate(data):
                    if isinstance(item, dict):
                        req_id = item.get('id', f"REQ{i+1:03d}")
                        req_text = item.get('text', item.get('requirement', item.get('description', "")))
                        priority = item.get('priority', "Medium")
                        category = item.get('category', "Functional")
                        
                        if req_text:
                            requirement = Requirement(
                                id=req_id,
                                text=req_text,
                                priority=priority,
                                category=category,
                                source=RequirementSource.JSON_FILE,
                                line_number=i+1
                            )
                            requirements.append(requirement)
            
            elif isinstance(data, dict):
                # Object with requirements key or direct requirement objects
                reqs_data = data.get('requirements', data)
                
                if isinstance(reqs_data, list):
                    for i, item in enumerate(reqs_data):
                        req_id = item.get('id', f"REQ{i+1:03d}")
                        req_text = item.get('text', item.get('requirement', ""))
                        priority = item.get('priority', "Medium")
                        category = item.get('category', "Functional")
                        
                        if req_text:
                            requirement = Requirement(
                                id=req_id,
                                text=req_text,
                                priority=priority,
                                category=category,
                                source=RequirementSource.JSON_FILE,
                                line_number=i+1
                            )
                            requirements.append(requirement)
                            
        except Exception as e:
            print(f"Error processing JSON file: {e}")
        
        return requirements
    
    @staticmethod
    def _process_text_file(file_path: str) -> List[Requirement]:
        """Process text file containing requirements"""
        requirements = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            
            lines = content.split('\n')
            req_counter = 1
            
            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue
                
                # Check if line starts with requirement identifier
                req_match = re.match(r'^(REQ[_-]?\d+|R\d+|\d+\.|\w+\d+):?\s*(.+)', line, re.IGNORECASE)
                if req_match:
                    req_id = req_match.group(1).replace(':', '')
                    req_text = req_match.group(2)
                    
                    requirement = Requirement(
                        id=req_id,
                        text=req_text,
                        priority="Medium",
                        category="Functional",
                        source=RequirementSource.TEXT_FILE,
                        line_number=line_num
                    )
                    requirements.append(requirement)
                    
                elif line.startswith(('-', '*', '•')):
                    # Bullet point format
                    req_text = line[1:].strip()
                    req_id = f"REQ{req_counter:03d}"
                    req_counter += 1
                    
                    requirement = Requirement(
                        id=req_id,
                        text=req_text,
                        priority="Medium",
                        category="Functional",
                        source=RequirementSource.TEXT_FILE,
                        line_number=line_num
                    )
                    requirements.append(requirement)
                    
                elif len(line) > 20:  # Assume longer lines are requirements
                    req_id = f"REQ{req_counter:03d}"
                    req_counter += 1
                    
                    requirement = Requirement(
                        id=req_id,
                        text=line,
                        priority="Medium",
                        category="Functional",
                        source=RequirementSource.TEXT_FILE,
                        line_number=line_num
                    )
                    requirements.append(requirement)
                    
        except Exception as e:
            print(f"Error processing text file: {e}")
        
        return requirements
    
    @staticmethod
    def _process_word_file(file_path: str) -> List[Requirement]:
        """Process Word document containing requirements"""
        requirements = []
        
        try:
            import docx
            
            doc = docx.Document(file_path)
            req_counter = 1
            
            for para_num, paragraph in enumerate(doc.paragraphs, 1):
                text = paragraph.text.strip()
                if not text or len(text) < 10:
                    continue
                
                # Check if paragraph starts with requirement identifier
                req_match = re.match(r'^(REQ[_-]?\d+|R\d+|\d+\.|\w+\d+):?\s*(.+)', text, re.IGNORECASE)
                if req_match:
                    req_id = req_match.group(1).replace(':', '')
                    req_text = req_match.group(2)
                else:
                    req_id = f"REQ{req_counter:03d}"
                    req_text = text
                    req_counter += 1
                
                # Check if this looks like a requirement (contains "shall", "must", "will", etc.)
                if any(keyword in text.lower() for keyword in ['shall', 'must', 'will', 'should', 'require']):
                    requirement = Requirement(
                        id=req_id,
                        text=req_text,
                        priority="Medium",
                        category="Functional",
                        source=RequirementSource.WORD_FILE,
                        line_number=para_num
                    )
                    requirements.append(requirement)
                    
        except ImportError:
            print("python-docx not installed. Install with: pip install python-docx")
        except Exception as e:
            print(f"Error processing Word file: {e}")
        
        return requirements

class AgentState(TypedDict):
    """State for the LangGraph agent"""
    messages: Annotated[List[BaseMessage], operator.add]
    requirements: List[Requirement]
    parsed_requirements: List[Dict[str, Any]]
    test_cases: List[TestCase]
    current_requirement: Optional[Dict[str, Any]]
    analysis_complete: bool
    generation_complete: bool
    validation_complete: bool
    optimization_complete: bool
    next_action: str
    file_sources: List[str]
    statistics: Dict[str, Any]

class AgenticTestCaseGenerator:
    """Fully agentic test case generator using LangGraph and OpenAI"""
    
    def __init__(self, llm_model: str = "gpt-4o", api_key: Optional[str] = None):
        self.llm = ChatOpenAI(
            model=llm_model,
            api_key=api_key or OPENAI_API_KEY,
            temperature=0.1
        )
        self.graph = self._build_graph()
        self.test_case_counter = 1
        
        # Advanced AI prompts for requirement analysis
        self.requirement_analysis_prompt = ChatPromptTemplate.from_messages([
            ("system", """You are an expert test analyst and requirement engineer. Analyze the given requirement comprehensively and extract:

1. Requirement Type: Classify as functional, non-functional, security, performance, usability, or integration
2. Functional Aspects: What the system should do
3. Non-functional Aspects: Performance, security, usability considerations
4. Input/Output Specifications: Expected inputs and outputs
5. Business Rules: Rules and constraints that apply
6. Error Conditions: What can go wrong and edge cases
7. Integration Points: External systems or components involved
8. Risk Assessment: High/Medium/Low based on complexity and criticality
9. Testability: How easily this can be tested
10. Boundary Conditions: Edge cases and limits
11. Security Considerations: Security implications
12. Performance Considerations: Performance requirements or implications

Return your analysis as a structured JSON object with clear categorization."""),
            ("human", "Requirement ID: {req_id}\nRequirement Text: {requirement}")
        ])
        
        # AI prompt for test case generation
        self.test_case_generation_prompt = ChatPromptTemplate.from_messages([
            ("system", """You are an expert test case designer. Based on the requirement analysis provided, generate comprehensive test cases that cover:

1. Positive Test Cases: Happy path scenarios (2-3 cases)
2. Negative Test Cases: Error scenarios and invalid inputs (2-4 cases)
3. Boundary Test Cases: Edge cases and limits (1-3 cases)
4. Security Test Cases: If security implications exist (1-2 cases)
5. Performance Test Cases: If performance requirements exist (1-2 cases)
6. Integration Test Cases: If integration points exist (1-2 cases)

For each test case, provide:
- Unique ID following pattern: {req_id}_[TYPE]_001
- Clear, descriptive title
- Detailed description
- Test type (positive, negative, boundary, security, performance, integration)
- Comprehensive preconditions
- Step-by-step test steps (be specific)
- Clear expected results
- Priority (High/Medium/Low)
- Risk level assessment
- Estimated execution time in minutes
- Relevant tags
- Test data requirements

Generate test cases as a JSON array. Be thorough and ensure comprehensive coverage."""),
            ("human", "Requirement Analysis: {analysis}\n\nGenerate comprehensive test cases for Requirement: {req_id}")
        ])
        
        # AI prompt for test case validation and enhancement
        self.validation_prompt = ChatPromptTemplate.from_messages([
            ("system", """You are a senior test manager reviewing test cases for quality and completeness. 

Analyze the provided test cases and:
1. Identify gaps in coverage
2. Suggest improvements to test steps
3. Enhance expected results for clarity
4. Optimize test case priorities
5. Identify missing edge cases
6. Suggest automation possibilities
7. Recommend risk-based prioritization

Return suggestions as structured JSON with specific recommendations."""),
            ("human", "Test Cases to Review: {test_cases}\n\nOriginal Requirement: {requirement}")
        ])
    
    def _build_graph(self) -> StateGraph:
        """Build the LangGraph workflow"""
        workflow = StateGraph(AgentState)
        
        # Add agentic nodes
        workflow.add_node("analyze_requirements", self._analyze_requirements_ai)
        workflow.add_node("generate_test_cases", self._generate_test_cases_ai)
        workflow.add_node("validate_test_cases", self._validate_test_cases_ai)
        workflow.add_node("optimize_test_suite", self._optimize_test_suite_ai)
        workflow.add_node("finalize_output", self._finalize_output)
        
        # Define workflow edges
        workflow.add_edge("analyze_requirements", "generate_test_cases")
        workflow.add_edge("generate_test_cases", "validate_test_cases")
        workflow.add_edge("validate_test_cases", "optimize_test_suite")
        workflow.add_edge("optimize_test_suite", "finalize_output")
        workflow.add_edge("finalize_output", END)
        
        # Set entry point
        workflow.set_entry_point("analyze_requirements")
        
        return workflow.compile()
    
    def _analyze_requirements_ai(self, state: AgentState) -> AgentState:
        """AI-powered requirement analysis"""
        print("🔍 AI Agent analyzing requirements...")
        
        parsed_requirements = []
        
        for req in state["requirements"]:
            print(f"   🤖 Analyzing {req.id} with AI...")
            
            try:
                # Use AI to analyze the requirement
                analysis_chain = self.requirement_analysis_prompt | self.llm
                
                response = analysis_chain.invoke({
                    "req_id": req.id,
                    "requirement": req.text
                })
                
                # Parse AI response
                analysis = self._parse_ai_requirement_analysis(response.content, req)
                parsed_requirements.append(analysis)
                
                print(f"   ✅ AI analysis complete for {req.id}")
                
            except Exception as e:
                print(f"   ⚠️ AI analysis failed for {req.id}: {e}")
                # Fallback to rule-based analysis
                analysis = self._fallback_requirement_analysis(req)
                parsed_requirements.append(analysis)
        
        state["parsed_requirements"] = parsed_requirements
        state["analysis_complete"] = True
        state["messages"].append(AIMessage(content=f"AI analyzed {len(parsed_requirements)} requirements"))
        
        return state
    
    def _generate_test_cases_ai(self, state: AgentState) -> AgentState:
        """AI-powered test case generation"""
        print("🚀 AI Agent generating test cases...")
        
        all_test_cases = []
        
        for req_analysis in state["parsed_requirements"]:
            print(f"   🤖 Generating test cases for {req_analysis['id']} with AI...")
            
            try:
                # Use AI to generate test cases
                generation_chain = self.test_case_generation_prompt | self.llm
                
                response = generation_chain.invoke({
                    "analysis": json.dumps(req_analysis, indent=2),
                    "req_id": req_analysis["id"]
                })
                
                # Parse AI-generated test cases
                test_cases = self._parse_ai_test_cases(response.content, req_analysis)
                
                if test_cases:
                    print(f"   ✅ AI generated {len(test_cases)} test cases for {req_analysis['id']}")
                    all_test_cases.extend(test_cases)
                else:
                    print(f"   ⚠️ AI generation failed, using enhanced templates for {req_analysis['id']}")
                    fallback_cases = self._generate_enhanced_template_test_cases(req_analysis)
                    all_test_cases.extend(fallback_cases)
                
            except Exception as e:
                print(f"   ❌ AI generation failed for {req_analysis['id']}: {e}")
                # Enhanced fallback
                fallback_cases = self._generate_enhanced_template_test_cases(req_analysis)
                all_test_cases.extend(fallback_cases)
        
        state["test_cases"] = all_test_cases
        state["generation_complete"] = True
        state["messages"].append(AIMessage(content=f"AI generated {len(all_test_cases)} test cases"))
        
        return state
    
    def _validate_test_cases_ai(self, state: AgentState) -> AgentState:
        """AI-powered test case validation and enhancement"""
        print("✅ AI Agent validating and enhancing test cases...")
        
        # Group test cases by requirement for validation
        test_cases_by_req = {}
        for tc in state["test_cases"]:
            req_id = tc.requirement_id
            if req_id not in test_cases_by_req:
                test_cases_by_req[req_id] = []
            test_cases_by_req[req_id].append(tc)
        
        enhanced_test_cases = []
        
        for req_id, test_cases in test_cases_by_req.items():
            print(f"   🤖 AI validating test cases for {req_id}...")
            
            try:
                # Find the original requirement
                original_req = None
                for req in state["requirements"]:
                    if req.id == req_id:
                        original_req = req
                        break
                
                if original_req:
                    # Use AI to validate and enhance
                    validation_chain = self.validation_prompt | self.llm
                    
                    test_cases_json = json.dumps([asdict(tc) for tc in test_cases], indent=2, default=str)
                    
                    response = validation_chain.invoke({
                        "test_cases": test_cases_json,
                        "requirement": original_req.text
                    })
                    
                    # Apply AI suggestions and enhance test cases
                    enhanced_cases = self._apply_ai_validation_suggestions(test_cases, response.content)
                    enhanced_test_cases.extend(enhanced_cases)
                    
                    print(f"   ✅ AI validation complete for {req_id}")
                else:
                    enhanced_test_cases.extend(test_cases)
                    
            except Exception as e:
                print(f"   ⚠️ AI validation failed for {req_id}: {e}")
                # Apply rule-based validation
                validated_cases = [self._validate_single_test_case(tc) for tc in test_cases]
                enhanced_test_cases.extend(validated_cases)
        
        state["test_cases"] = enhanced_test_cases
        state["validation_complete"] = True
        state["messages"].append(AIMessage(content=f"AI validated and enhanced {len(enhanced_test_cases)} test cases"))
        
        return state
    
    def _optimize_test_suite_ai(self, state: AgentState) -> AgentState:
        """AI-powered test suite optimization"""
        print("⚡ AI Agent optimizing test suite...")
        
        # Remove duplicates and optimize
        unique_cases = self._remove_duplicates_intelligent(state["test_cases"])
        
        # AI-powered prioritization
        prioritized_cases = self._ai_prioritize_test_cases(unique_cases)
        
        # Add comprehensive traceability
        traceable_cases = self._add_enhanced_traceability(prioritized_cases, state["parsed_requirements"])
        
        # Generate coverage metrics
        coverage_metrics = self._calculate_coverage_metrics(traceable_cases, state["requirements"])
        
        state["test_cases"] = traceable_cases
        state["optimization_complete"] = True
        state["statistics"] = coverage_metrics
        state["messages"].append(AIMessage(content=f"AI optimized test suite: {len(traceable_cases)} test cases"))
        
        return state
    
    def _finalize_output(self, state: AgentState) -> AgentState:
        """Finalize the agentic output"""
        print("🎯 Finalizing agentic test suite...")
        
        # Generate comprehensive statistics
        final_stats = self._generate_comprehensive_statistics(state["test_cases"], state["requirements"])
        state["statistics"].update(final_stats)
        
        # Generate quality assessment
        quality_metrics = self._assess_test_suite_quality(state["test_cases"])
        state["statistics"]["quality_metrics"] = quality_metrics
        
        summary = f"Agentic test suite complete: {len(state['test_cases'])} test cases with {quality_metrics['overall_score']}% quality score"
        state["messages"].append(AIMessage(content=summary))
        
        return state
    
    def _parse_ai_requirement_analysis(self, ai_response: str, req: Requirement) -> Dict[str, Any]:
        """Parse AI requirement analysis response"""
        try:
            # Extract JSON from AI response
            json_match = re.search(r'\{.*\}', ai_response, re.DOTALL)
            if json_match:
                analysis_data = json.loads(json_match.group())
            else:
                analysis_data = {}
        except:
            analysis_data = {}
        
        # Ensure all required fields with AI insights
        return {
            "id": req.id,
            "original_text": req.text,
            "requirement_type": analysis_data.get("requirement_type", "functional"),
            "functional_aspects": analysis_data.get("functional_aspects", []),
            "non_functional_aspects": analysis_data.get("non_functional_aspects", []),
            "business_rules": analysis_data.get("business_rules", []),
            "error_conditions": analysis_data.get("error_conditions", []),
            "integration_points": analysis_data.get("integration_points", []),
            "boundary_conditions": analysis_data.get("boundary_conditions", []),
            "security_considerations": analysis_data.get("security_considerations", {}),
            "performance_considerations": analysis_data.get("performance_considerations", {}),
            "risk_level": analysis_data.get("risk_assessment", "Medium"),
            "testability": analysis_data.get("testability", "High"),
            "complexity": analysis_data.get("complexity", "Medium"),
            "dependencies": analysis_data.get("dependencies", [])
        }
    
    def _fallback_requirement_analysis(self, req: Requirement) -> Dict[str, Any]:
        """Enhanced fallback analysis when AI fails"""
        req_text_lower = req.text.lower()
        
        # Determine requirement type
        if any(keyword in req_text_lower for keyword in ['security', 'authentication', 'authorization', 'encrypt']):
            req_type = "security"
        elif any(keyword in req_text_lower for keyword in ['performance', 'speed', 'response time', 'load']):
            req_type = "performance"
        elif any(keyword in req_text_lower for keyword in ['usability', 'user interface', 'user experience']):
            req_type = "usability"
        elif any(keyword in req_text_lower for keyword in ['integration', 'api', 'interface', 'connect']):
            req_type = "integration"
        else:
            req_type = "functional"
        
        return {
            "id": req.id,
            "original_text": req.text,
            "requirement_type": req_type,
            "functional_aspects": ["Core functionality implementation"],
            "non_functional_aspects": [],
            "business_rules": [],
            "error_conditions": ["Invalid input", "System unavailable"],
            "integration_points": [],
            "boundary_conditions": [],
            "security_considerations": {},
            "performance_considerations": {},
            "risk_level": "Medium",
            "testability": "High",
            "complexity": "Medium",
            "dependencies": []
        }
    
    def _parse_ai_test_cases(self, ai_response: str, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Parse AI-generated test cases"""
        test_cases = []
        
        try:
            # Extract JSON array from AI response
            json_match = re.search(r'\[.*\]', ai_response, re.DOTALL)
            if json_match:
                cases_data = json.loads(json_match.group())
                
                for case_data in cases_data:
                    try:
                        # Parse test type
                        test_type_str = case_data.get("test_type", "positive").lower()
                        test_type = self._parse_test_type(test_type_str)
                        
                        test_case = TestCase(
                            id=case_data.get("id", f"{req_analysis['id']}_AI_{self.test_case_counter:03d}"),
                            title=case_data.get("title", "AI Generated Test Case"),
                            description=case_data.get("description", "AI generated test case"),
                            test_type=test_type,
                            preconditions=case_data.get("preconditions", ["System is available"]),
                            test_steps=case_data.get("test_steps", ["Execute test"]),
                            expected_result=case_data.get("expected_result", "Expected behavior occurs"),
                            priority=case_data.get("priority", "Medium"),
                            tags=case_data.get("tags", [req_analysis['id'].lower()]),
                            risk_level=case_data.get("risk_level", "Medium"),
                            requirement_id=req_analysis['id'],
                            estimated_time=case_data.get("estimated_time", 30),
                            test_data=case_data.get("test_data", [])
                        )
                        test_cases.append(test_case)
                        self.test_case_counter += 1
                    except Exception as e:
                        print(f"Error parsing individual AI test case: {e}")
                        continue
        except Exception as e:
            print(f"Error parsing AI test cases: {e}")
        
        return test_cases
    
    def _parse_test_type(self, test_type_str: str) -> TestType:
        """Parse test type string to enum"""
        type_mapping = {
            "positive": TestType.POSITIVE,
            "negative": TestType.NEGATIVE,
            "boundary": TestType.BOUNDARY,
            "edge_case": TestType.EDGE_CASE,
            "security": TestType.SECURITY,
            "performance": TestType.PERFORMANCE,
            "integration": TestType.INTEGRATION,
            "usability": TestType.USABILITY
        }
        return type_mapping.get(test_type_str, TestType.POSITIVE)
    
    def _apply_ai_validation_suggestions(self, test_cases: List[TestCase], ai_suggestions: str) -> List[TestCase]:
        """Apply AI validation suggestions to enhance test cases"""
        # For now, apply basic validation and return enhanced cases
        # In a full implementation, this would parse AI suggestions and apply them
        enhanced_cases = []
        
        for tc in test_cases:
            enhanced_tc = self._validate_single_test_case(tc)
            enhanced_cases.append(enhanced_tc)
        
        return enhanced_cases
    
    def _validate_single_test_case(self, test_case: TestCase) -> TestCase:
        """Validate and enhance a single test case"""
        # Enhance title if too short
        if len(test_case.title) < 15:
            test_case.title = f"Verify {test_case.requirement_id} - {test_case.title}"
        
        # Ensure minimum test steps
        if len(test_case.test_steps) < 3:
            test_case.test_steps.extend([
                "Verify system state",
                "Validate results"
            ])
        
        # Enhance expected result if generic
        if test_case.expected_result in ["Expected behavior occurs", "System behaves as expected"]:
            test_case.expected_result = f"System successfully completes {test_case.test_type.value} test scenario without errors"
        
        return test_case
    
    def _remove_duplicates_intelligent(self, test_cases: List[TestCase]) -> List[TestCase]:
        """Intelligently remove duplicate test cases"""
        unique_cases = []
        seen_combinations = set()
        
        for case in test_cases:
            # Create a signature based on test type, requirement, and key steps
            signature = (
                case.requirement_id,
                case.test_type.value,
                case.title.lower().replace(" ", "")[:50]
            )
            
            if signature not in seen_combinations:
                seen_combinations.add(signature)
                unique_cases.append(case)
        
        return unique_cases
    
    def _ai_prioritize_test_cases(self, test_cases: List[TestCase]) -> List[TestCase]:
        """AI-powered test case prioritization"""
        # Priority scoring based on multiple factors
        def calculate_priority_score(tc):
            score = 0
            
            # Priority weight
            if tc.priority == "High":
                score += 10
            elif tc.priority == "Medium":
                score += 5
            
            # Risk weight
            if tc.risk_level == "High":
                score += 8
            elif tc.risk_level == "Medium":
                score += 4
            
            # Test type weight
            type_weights = {
                TestType.SECURITY: 9,
                TestType.POSITIVE: 8,
                TestType.NEGATIVE: 7,
                TestType.INTEGRATION: 6,
                TestType.BOUNDARY: 5,
                TestType.PERFORMANCE: 4,
                TestType.USABILITY: 3,
                TestType.EDGE_CASE: 2
            }
            score += type_weights.get(tc.test_type, 2)
            
            return score
        
        return sorted(test_cases, key=calculate_priority_score, reverse=True)
    
    def _add_enhanced_traceability(self, test_cases: List[TestCase], requirements: List[Dict[str, Any]]) -> List[TestCase]:
        """Add enhanced traceability information"""
        req_map = {req["id"]: req for req in requirements}
        
        for case in test_cases:
            # Add traceability tag
            if case.requirement_id in req_map:
                case.tags.append(f"traces_to_{case.requirement_id}")
                
                # Add requirement type tag
                req_type = req_map[case.requirement_id].get("requirement_type", "functional")
                case.tags.append(f"req_type_{req_type}")
        
        return test_cases
    
    def _calculate_coverage_metrics(self, test_cases: List[TestCase], requirements: List[Requirement]) -> Dict[str, Any]:
        """Calculate comprehensive coverage metrics"""
        total_reqs = len(requirements)
        covered_reqs = len(set(tc.requirement_id for tc in test_cases))
        
        # Calculate coverage by test type
        type_coverage = {}
        for test_type in TestType:
            type_count = sum(1 for tc in test_cases if tc.test_type == test_type)
            type_coverage[test_type.value] = type_count
        
        return {
            "requirement_coverage": {
                "total_requirements": total_reqs,
                "covered_requirements": covered_reqs,
                "coverage_percentage": round((covered_reqs / total_reqs) * 100, 2) if total_reqs > 0 else 0
            },
            "test_type_distribution": type_coverage,
            "total_test_cases": len(test_cases)
        }
    
    def _generate_comprehensive_statistics(self, test_cases: List[TestCase], requirements: List[Requirement]) -> Dict[str, Any]:
        """Generate comprehensive statistics"""
        return {
            "execution_estimates": {
                "total_minutes": sum(tc.estimated_time for tc in test_cases),
                "total_hours": round(sum(tc.estimated_time for tc in test_cases) / 60, 2),
                "high_priority_minutes": sum(tc.estimated_time for tc in test_cases if tc.priority == "High")
            },
            "automation_analysis": {
                "automatable_tests": sum(1 for tc in test_cases if tc.automation_feasible),
                "manual_tests": sum(1 for tc in test_cases if not tc.automation_feasible),
                "automation_percentage": round((sum(1 for tc in test_cases if tc.automation_feasible) / len(test_cases)) * 100, 2) if test_cases else 0
            }
        }
    
    def _assess_test_suite_quality(self, test_cases: List[TestCase]) -> Dict[str, Any]:
        """Assess overall test suite quality"""
        if not test_cases:
            return {"overall_score": 0}
        
        # Quality factors
        has_positive = any(tc.test_type == TestType.POSITIVE for tc in test_cases)
        has_negative = any(tc.test_type == TestType.NEGATIVE for tc in test_cases)
        has_boundary = any(tc.test_type == TestType.BOUNDARY for tc in test_cases)
        has_security = any(tc.test_type == TestType.SECURITY for tc in test_cases)
        
        avg_steps = sum(len(tc.test_steps) for tc in test_cases) / len(test_cases)
        detailed_cases = sum(1 for tc in test_cases if len(tc.test_steps) >= 4) / len(test_cases)
        
        # Calculate quality score
        score = 0
        if has_positive: score += 20
        if has_negative: score += 20
        if has_boundary: score += 15
        if has_security: score += 15
        if avg_steps >= 4: score += 15
        if detailed_cases >= 0.7: score += 15
        
        return {
            "overall_score": min(100, score),
            "coverage_types": {
                "positive": has_positive,
                "negative": has_negative,
                "boundary": has_boundary,
                "security": has_security
            },
            "detail_quality": {
                "avg_steps_per_test": round(avg_steps, 1),
                "detailed_tests_percentage": round(detailed_cases * 100, 1)
            }
        }
    
    def _generate_enhanced_template_test_cases(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Enhanced template-based test case generation as fallback"""
        test_cases = []
        req_id = req_analysis["id"]
        req_text = req_analysis["original_text"].lower()
        req_type = req_analysis.get("requirement_type", "functional")
        
        # Generate based on requirement type
        if "password" in req_text:
            test_cases.extend(self._generate_password_test_cases_enhanced(req_analysis))
        elif "email" in req_text:
            test_cases.extend(self._generate_email_test_cases_enhanced(req_analysis))
        elif "authentication" in req_text or "login" in req_text:
            test_cases.extend(self._generate_auth_test_cases_enhanced(req_analysis))
        elif req_type == "security":
            test_cases.extend(self._generate_security_test_cases_enhanced(req_analysis))
        elif req_type == "performance":
            test_cases.extend(self._generate_performance_test_cases_enhanced(req_analysis))
        else:
            test_cases.extend(self._generate_generic_test_cases_enhanced(req_analysis))
        
        return test_cases
    
    def _generate_password_test_cases_enhanced(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate enhanced password test cases"""
        req_id = req_analysis["id"]
        test_cases = []
        
        # Positive cases
        positive_case = TestCase(
            id=f"{req_id}_POSITIVE_{self.test_case_counter:03d}",
            title=f"Verify {req_id} - Strong password acceptance",
            description="Test that a password meeting all security requirements is accepted",
            test_type=TestType.POSITIVE,
            preconditions=["Password input field is accessible", "Password requirements are displayed"],
            test_steps=[
                "Navigate to password input field",
                "Enter password meeting all criteria: 'SecurePass123!'",
                "Submit password for validation",
                "Verify acceptance without errors",
                "Confirm password strength indicator shows 'Strong'"
            ],
            expected_result="Password is accepted, validation passes, and strength indicator shows strong",
            priority="High",
            tags=["password", "validation", "positive", "security"],
            risk_level="High",
            requirement_id=req_id,
            estimated_time=25,
            test_data=["SecurePass123!", "ValidP@ssw0rd", "Strong&Pass2024"]
        )
        test_cases.append(positive_case)
        self.test_case_counter += 1
        
        # Negative cases with detailed scenarios
        negative_scenarios = [
            ("Too short", "Test1!", "Password must be at least 8 characters"),
            ("Missing uppercase", "testpass123!", "Password must contain uppercase letter"),
            ("Missing digit", "TestPassword!", "Password must contain at least one digit"),
            ("Missing special char", "TestPassword123", "Password must contain special character"),
            ("Too long", "A" * 130 + "1!", "Password exceeds maximum length")
        ]
        
        for scenario, test_pass, expected_error in negative_scenarios:
            negative_case = TestCase(
                id=f"{req_id}_NEGATIVE_{self.test_case_counter:03d}",
                title=f"Verify {req_id} - {scenario}",
                description=f"Test password validation rejects password that is {scenario.lower()}",
                test_type=TestType.NEGATIVE,
                preconditions=["Password input field is accessible"],
                test_steps=[
                    "Navigate to password input field",
                    f"Enter invalid password: '{test_pass}'",
                    "Submit password for validation",
                    "Verify validation error is displayed",
                    "Confirm password is not accepted"
                ],
                expected_result=f"Validation fails with error: '{expected_error}'",
                priority="High",
                tags=["password", "validation", "negative", "security"],
                risk_level="Medium",
                requirement_id=req_id,
                estimated_time=20,
                test_data=[test_pass]
            )
            test_cases.append(negative_case)
            self.test_case_counter += 1
        
        return test_cases
    
    def _generate_email_test_cases_enhanced(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate enhanced email validation test cases"""
        req_id = req_analysis["id"]
        test_cases = []
        
        # Positive cases with various valid formats
        valid_emails = [
            ("Standard format", "user@example.com"),
            ("With subdomain", "user@mail.example.com"),
            ("With plus addressing", "user+tag@example.com"),
            ("International domain", "user@example.co.uk")
        ]
        
        for desc, email in valid_emails:
            positive_case = TestCase(
                id=f"{req_id}_POSITIVE_{self.test_case_counter:03d}",
                title=f"Verify {req_id} - {desc}",
                description=f"Test email validation accepts valid format: {email}",
                test_type=TestType.POSITIVE,
                preconditions=["Email input field is accessible"],
                test_steps=[
                    "Navigate to email input field",
                    f"Enter valid email: '{email}'",
                    "Submit form or trigger validation",
                    "Verify email is accepted",
                    "Confirm no validation errors"
                ],
                expected_result="Email is accepted as valid without errors",
                priority="High",
                tags=["email", "validation", "positive"],
                risk_level="Low",
                requirement_id=req_id,
                estimated_time=15,
                test_data=[email]
            )
            test_cases.append(positive_case)
            self.test_case_counter += 1
        
        return test_cases
    
    def _generate_auth_test_cases_enhanced(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate enhanced authentication test cases"""
        req_id = req_analysis["id"]
        test_cases = []
        
        # Comprehensive positive authentication test
        positive_case = TestCase(
            id=f"{req_id}_POSITIVE_{self.test_case_counter:03d}",
            title=f"Verify {req_id} - Successful authentication flow",
            description="Test complete successful authentication with valid credentials",
            test_type=TestType.POSITIVE,
            preconditions=["Valid user account exists", "Login page is accessible", "Network connectivity available"],
            test_steps=[
                "Navigate to login page",
                "Verify login form is displayed correctly",
                "Enter valid username in username field",
                "Enter valid password in password field",
                "Click login/submit button",
                "Verify successful authentication",
                "Confirm redirect to dashboard/home page",
                "Verify user session is established"
            ],
            expected_result="User successfully authenticates and is redirected to authorized area with active session",
            priority="High",
            tags=["authentication", "login", "positive", "security"],
            risk_level="High",
            requirement_id=req_id,
            estimated_time=35
        )
        test_cases.append(positive_case)
        self.test_case_counter += 1
        
        return test_cases
    
    def _generate_security_test_cases_enhanced(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate enhanced security test cases"""
        req_id = req_analysis["id"]
        test_cases = []
        
        # Security vulnerability test
        security_case = TestCase(
            id=f"{req_id}_SECURITY_{self.test_case_counter:03d}",
            title=f"Verify {req_id} - Security vulnerability assessment",
            description="Test system security against common vulnerabilities",
            test_type=TestType.SECURITY,
            preconditions=["Security testing tools available", "Test environment configured"],
            test_steps=[
                "Identify input fields and parameters",
                "Test SQL injection with malicious payloads",
                "Test XSS with script injection attempts",
                "Verify input sanitization and validation",
                "Test authentication bypass attempts",
                "Verify secure data transmission (HTTPS)",
                "Test session management security"
            ],
            expected_result="System properly validates inputs, prevents injections, and maintains security standards",
            priority="High",
            tags=["security", "vulnerability", "penetration"],
            risk_level="High",
            requirement_id=req_id,
            estimated_time=90
        )
        test_cases.append(security_case)
        self.test_case_counter += 1
        
        return test_cases
    
    def _generate_performance_test_cases_enhanced(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate enhanced performance test cases"""
        req_id = req_analysis["id"]
        test_cases = []
        
        # Load performance test
        perf_case = TestCase(
            id=f"{req_id}_PERFORMANCE_{self.test_case_counter:03d}",
            title=f"Verify {req_id} - Performance under load",
            description="Test system performance under expected load conditions",
            test_type=TestType.PERFORMANCE,
            preconditions=["Performance testing environment ready", "Load testing tools configured"],
            test_steps=[
                "Configure load testing parameters",
                "Set up performance monitoring",
                "Execute functionality with baseline load",
                "Gradually increase user load to expected levels",
                "Monitor response times and resource usage",
                "Identify performance bottlenecks",
                "Verify system meets performance requirements"
            ],
            expected_result="System maintains acceptable performance under expected load conditions",
            priority="High",
            tags=["performance", "load", "scalability"],
            risk_level="Medium",
            requirement_id=req_id,
            estimated_time=120
        )
        test_cases.append(perf_case)
        self.test_case_counter += 1
        
        return test_cases
    
    def _generate_generic_test_cases_enhanced(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate enhanced generic test cases"""
        req_id = req_analysis["id"]
        test_cases = []
        
        # Comprehensive positive test
        positive_case = TestCase(
            id=f"{req_id}_POSITIVE_{self.test_case_counter:03d}",
            title=f"Verify {req_id} - Functional requirement implementation",
            description=f"Test successful implementation of requirement: {req_analysis['original_text'][:100]}...",
            test_type=TestType.POSITIVE,
            preconditions=["System is operational", "Required test data available", "User has appropriate permissions"],
            test_steps=[
                "Verify system prerequisites",
                "Prepare valid test data",
                "Execute the required functionality",
                "Verify expected behavior occurs",
                "Validate output/results",
                "Confirm system state is correct"
            ],
            expected_result="System successfully implements the requirement as specified",
            priority="High",
            tags=["functional", "positive", req_analysis.get("requirement_type", "generic")],
            risk_level=req_analysis.get("risk_level", "Medium"),
            requirement_id=req_id,
            estimated_time=40
        )
        test_cases.append(positive_case)
        self.test_case_counter += 1
        
        # Comprehensive negative test
        negative_case = TestCase(
            id=f"{req_id}_NEGATIVE_{self.test_case_counter:03d}",
            title=f"Verify {req_id} - Error handling and validation",
            description="Test system error handling with invalid inputs and conditions",
            test_type=TestType.NEGATIVE,
            preconditions=["System is operational", "Error handling mechanisms in place"],
            test_steps=[
                "Prepare invalid or malformed test data",
                "Attempt to execute functionality with invalid inputs",
                "Verify appropriate error handling",
                "Confirm error messages are user-friendly",
                "Verify system remains stable",
                "Confirm proper logging of errors"
            ],
            expected_result="System gracefully handles errors with appropriate messages and maintains stability",
            priority="Medium",
            tags=["functional", "negative", "error_handling"],
            risk_level="Medium",
            requirement_id=req_id,
            estimated_time=35
        )
        test_cases.append(negative_case)
        self.test_case_counter += 1
        
        return test_cases
    
    # Main interface methods
    def generate_test_cases_from_files(self, file_paths: List[str]) -> List[TestCase]:
        """Generate test cases from uploaded requirement files using AI agents"""
        all_requirements = []
        
        print(f"📁 Processing {len(file_paths)} requirement files...")
        
        for file_path in file_paths:
            print(f"   📄 Processing {file_path}...")
            requirements = FileProcessor.process_file(file_path)
            
            if requirements:
                print(f"   ✅ Found {len(requirements)} requirements in {file_path}")
                all_requirements.extend(requirements)
            else:
                print(f"   ⚠️ No requirements found in {file_path}")
        
        if not all_requirements:
            print("❌ No requirements found in any files")
            return []
        
        print(f"📋 Total requirements loaded: {len(all_requirements)}")
        
        return self._generate_test_cases_from_requirements(all_requirements, file_paths)
    
    def generate_test_cases_from_manual_input(self, requirements: List[tuple]) -> List[TestCase]:
        """Generate test cases from manually entered requirements using AI agents"""
        req_objects = []
        for req_id, req_text in requirements:
            req_obj = Requirement(
                id=req_id,
                text=req_text,
                source=RequirementSource.MANUAL
            )
            req_objects.append(req_obj)
        
        return self._generate_test_cases_from_requirements(req_objects, [])
    
    def _generate_test_cases_from_requirements(self, requirements: List[Requirement], file_sources: List[str]) -> List[TestCase]:
        """Internal method to generate test cases using AI agents"""
        # Initial state for the AI agent workflow
        initial_state = {
            "messages": [HumanMessage(content="Starting agentic test case generation")],
            "requirements": requirements,
            "parsed_requirements": [],
            "test_cases": [],
            "current_requirement": None,
            "analysis_complete": False,
            "generation_complete": False,
            "validation_complete": False,
            "optimization_complete": False,
            "next_action": "analyze",
            "file_sources": file_sources,
            "statistics": {}
        }
        
        # Run the AI agent workflow
        final_state = self.graph.invoke(initial_state)
        
        return final_state["test_cases"]
    
    def export_to_json(self, test_cases: List[TestCase], filename: str = "test_cases.json"):
        """Export test cases to JSON"""
        cases_dict = []
        for case in test_cases:
            case_dict = asdict(case)
            case_dict["test_type"] = case.test_type.value
            cases_dict.append(case_dict)
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(cases_dict, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Exported {len(test_cases)} test cases to {filename}")
    
    def export_to_csv(self, test_cases: List[TestCase], filename: str = "test_cases.csv"):
        """Export test cases to CSV format"""
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Test ID', 'Title', 'Description', 'Type', 'Priority', 
                         'Preconditions', 'Test Steps', 'Expected Result', 'Tags',
                         'Risk Level', 'Requirement ID', 'Automation Feasible',
                         'Estimated Time (min)', 'Test Data']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for tc in test_cases:
                writer.writerow({
                    'Test ID': tc.id,
                    'Title': tc.title,
                    'Description': tc.description,
                    'Type': tc.test_type.value,
                    'Priority': tc.priority,
                    'Preconditions': '; '.join(tc.preconditions),
                    'Test Steps': '; '.join(tc.test_steps),
                    'Expected Result': tc.expected_result,
                    'Tags': ', '.join(tc.tags),
                    'Risk Level': tc.risk_level,
                    'Requirement ID': tc.requirement_id,
                    'Automation Feasible': tc.automation_feasible,
                    'Estimated Time (min)': tc.estimated_time,
                    'Test Data': '; '.join(tc.test_data)
                })
        
        print(f"✅ Exported {len(test_cases)} test cases to {filename}")
    
    def print_test_cases(self, test_cases: List[TestCase]):
        """Print test cases in a readable format"""
        for tc in test_cases:
            print(f"\n{'='*60}")
            print(f"Test ID: {tc.id}")
            print(f"Title: {tc.title}")
            print(f"Type: {tc.test_type.value.upper()}")
            print(f"Priority: {tc.priority}")
            print(f"Risk Level: {tc.risk_level}")
            print(f"Estimated Time: {tc.estimated_time} minutes")
            print(f"Description: {tc.description}")
            print(f"\nPreconditions:")
            for i, precond in enumerate(tc.preconditions, 1):
                print(f"  {i}. {precond}")
            print(f"\nTest Steps:")
            for i, step in enumerate(tc.test_steps, 1):
                print(f"  {i}. {step}")
            print(f"\nExpected Result: {tc.expected_result}")
            print(f"Tags: {', '.join(tc.tags)}")
            print(f"Automation Feasible: {tc.automation_feasible}")
            if tc.test_data:
                print(f"Test Data: {'; '.join(tc.test_data)}")

# Demo/Main function
def main():
    """Demonstration of the Agentic Test Case Generator with AI"""
    
    def create_sample_files():
        """Create sample requirement files for demonstration"""
        
        # Sample CSV file
        csv_content = """ID,Requirement,Priority,Category
REQ001,"The system shall validate user email addresses using RFC 5322 standard and reject invalid formats with appropriate error messages",High,Authentication
REQ002,"User passwords must be between 8 and 128 characters long and contain at least one uppercase letter, one lowercase letter, one digit, and one special character",High,Security
REQ003,"The system shall implement rate limiting to prevent brute force attacks, allowing maximum 5 login attempts per IP address within 15 minutes",High,Security
REQ004,"When a user enters invalid credentials 3 consecutive times, the account shall be temporarily locked for 30 minutes",High,Security"""
        
        with open('comprehensive_requirements.csv', 'w', encoding='utf-8') as f:
            f.write(csv_content)
        
        return ['comprehensive_requirements.csv']
    
    print("🤖 Starting Agentic Test Case Generation with AI!")
    print("🧠 Using OpenAI GPT-4 for intelligent test case generation")
    print("=" * 80)
    
    try:
        generator = AgenticTestCaseGenerator()
        
        print("🔧 Creating sample requirement files...")
        sample_files = create_sample_files()
        print(f"   ✅ Created: {', '.join(sample_files)}")
        
        print(f"\n📂 Processing requirement files with AI agents...")
        test_cases = generator.generate_test_cases_from_files(sample_files)
        
        # Display results
        print(f"\n🎉 AI Generated {len(test_cases)} comprehensive test cases!")
        
        # Summary by requirement
        print(f"\n📊 Test cases by requirement:")
        by_req = {}
        for tc in test_cases:
            req_id = tc.requirement_id
            by_req[req_id] = by_req.get(req_id, 0) + 1
        
        for req_id, count in sorted(by_req.items()):
            print(f"   {req_id}: {count} test cases")
        
        # Summary by test type
        print(f"\n🔍 Test cases by type:")
        by_type = {}
        for tc in test_cases:
            test_type = tc.test_type.value
            by_type[test_type] = by_type.get(test_type, 0) + 1
        
        for test_type, count in sorted(by_type.items()):
            print(f"   {test_type.title()}: {count} test cases")
        
        # Show detailed test cases for one requirement
        print(f"\n🔐 Sample: REQ002 (Password) test cases:")
        req002_cases = [tc for tc in test_cases if tc.requirement_id == "REQ002"]
        for tc in req002_cases[:3]:  # Show first 3
            print(f"   - {tc.id}: {tc.title}")
            print(f"     Type: {tc.test_type.value}, Priority: {tc.priority}, Time: {tc.estimated_time}min")
        
        # Export results
        print(f"\n💾 Exporting AI-generated test cases...")
        generator.export_to_json(test_cases, "comprehensive_test_cases.json")
        print(f"   ✅ Exported to JSON: comprehensive_test_cases.json")
        
        generator.export_to_csv(test_cases, "comprehensive_test_cases.csv")
        print(f"   ✅ Exported to CSV: comprehensive_test_cases.csv")
        
        # Usage examples
        print(f"\n📚 Usage Examples:")
        print(f"   File processing: generator.generate_test_cases_from_files(['requirements.csv'])")
        print(f"   Manual input: generator.generate_test_cases_from_manual_input([('REQ001', 'requirement text')])")
        
        # Clean up
        import os
        for file in sample_files:
            try:
                os.remove(file)
            except:
                pass
        print(f"\n🧹 Cleaned up sample files")
        
        print(f"\n✨ Agentic test case generation complete!")
        print(f"🎯 Generated comprehensive test suite with AI-powered analysis and optimization")
        
    except Exception as e:
        print(f"❌ Error during agentic test case generation: {e}")
        print("💡 Make sure you have set OPENAI_API_KEY in your environment variables")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()
