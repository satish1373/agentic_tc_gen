import re
import json
import csv
import io
from typing import List, Dict, Any, Optional, TypedDict, Annotated, Union
from dataclasses import dataclass, asdict
from enum import Enum
import operator

#LangGraph imports - commented out for demo
from langgraph.graph import StateGraph, END
#from langgraph.prebuilt import ToolExecutor
import langgraph.prebuilt
from langchain_core.messages import BaseMessage, HumanMessage, AIMessage
from langchain_core.tools import Tool
from langchain_core.prompts import ChatPromptTemplate
from langchain_openai import ChatOpenAI

from dotenv import load_dotenv
import os

# Load environment variables from .env file
load_dotenv()

# Access the key
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY")

# Mock classes for demo purposes
class BaseMessage:
    def __init__(self, content: str):
        self.content = content

class HumanMessage(BaseMessage):
    pass

class AIMessage(BaseMessage):
    pass

class StateGraph:
    def __init__(self, state_class):
        self.state_class = state_class
        self.nodes = {}
        self.edges = []
        self.entry_point = None
    
    def add_node(self, name: str, func):
        self.nodes[name] = func
    
    def add_edge(self, from_node: str, to_node: str):
        self.edges.append((from_node, to_node))
    
    def set_entry_point(self, node: str):
        self.entry_point = node
    
    def compile(self):
        return MockGraph(self.nodes, self.edges, self.entry_point)

class MockGraph:
    def __init__(self, nodes, edges, entry_point):
        self.nodes = nodes
        self.edges = edges
        self.entry_point = entry_point
    
    def invoke(self, initial_state):
        state = initial_state
        current_node = self.entry_point
        
        # Simple execution flow
        execution_order = ["analyze_requirements", "generate_test_cases", 
                          "validate_test_cases", "optimize_test_suite", "finalize_output"]
        
        for node_name in execution_order:
            if node_name in self.nodes:
                state = self.nodes[node_name](state)
        
        return state

END = "END"

class TestType(Enum):
    POSITIVE = "positive"
    NEGATIVE = "negative"
    BOUNDARY = "boundary"
    EDGE_CASE = "edge_case"
    INTEGRATION = "integration"
    SECURITY = "security"

class RequirementSource(Enum):
    MANUAL = "manual"
    CSV_FILE = "csv_file"
    EXCEL_FILE = "excel_file"
    JSON_FILE = "json_file"
    TEXT_FILE = "text_file"
    WORD_FILE = "word_file"

@dataclass
class Requirement:
    """Represents a single requirement"""
    id: str
    text: str
    priority: str = "Medium"
    category: str = "Functional"
    source: RequirementSource = RequirementSource.MANUAL
    line_number: Optional[int] = None
    
@dataclass
class TestCase:
    """Represents a single test case"""
    id: str
    title: str
    description: str
    test_type: TestType
    preconditions: List[str]
    test_steps: List[str]
    expected_result: str
    priority: str = "Medium"
    tags: List[str] = None
    risk_level: str = "Medium"
    automation_feasible: bool = True
    requirement_id: str = ""
    
    def __post_init__(self):
        if self.tags is None:
            self.tags = []

class FileProcessor:
    """Handles processing of various file formats containing requirements"""
    
    @staticmethod
    def process_file(file_path: str) -> List[Requirement]:
        """Process a file and extract requirements based on file type"""
        try:
            file_extension = file_path.lower().split('.')[-1]
            
            if file_extension == 'csv':
                return FileProcessor._process_csv_file(file_path)
            elif file_extension in ['xlsx', 'xls']:
                return FileProcessor._process_excel_file(file_path)
            elif file_extension == 'json':
                return FileProcessor._process_json_file(file_path)
            elif file_extension == 'txt':
                return FileProcessor._process_text_file(file_path)
            elif file_extension in ['docx', 'doc']:
                return FileProcessor._process_word_file(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
                
        except Exception as e:
            print(f"Error processing file {file_path}: {e}")
            return []
    
    @staticmethod
    def _process_csv_file(file_path: str) -> List[Requirement]:
        """Process CSV file containing requirements"""
        requirements = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
        except Exception as e:
            print(f"Error reading CSV file: {e}")
            return []
        
        # Parse CSV content
        csv_reader = csv.DictReader(io.StringIO(content))
        
        for row_num, row in enumerate(csv_reader, 1):
            # Try different possible column names
            req_id = (row.get('ID') or row.get('id') or 
                     row.get('Requirement ID') or row.get('req_id') or 
                     f"REQ{row_num:03d}")
            
            req_text = (row.get('Requirement') or row.get('requirement') or 
                       row.get('Description') or row.get('description') or 
                       row.get('Text') or row.get('text') or "")
            
            priority = (row.get('Priority') or row.get('priority') or "Medium")
            category = (row.get('Category') or row.get('category') or "Functional")
            
            if req_text.strip():
                requirement = Requirement(
                    id=req_id,
                    text=req_text.strip(),
                    priority=priority,
                    category=category,
                    source=RequirementSource.CSV_FILE,
                    line_number=row_num
                )
                requirements.append(requirement)
        
        return requirements
    
    @staticmethod
    def _process_excel_file(file_path: str) -> List[Requirement]:
        """Process Excel file containing requirements"""
        requirements = []
        
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path)
            worksheet = workbook.active
            
            # Find header row
            headers = {}
            for col in range(1, worksheet.max_column + 1):
                header_value = worksheet.cell(1, col).value
                if header_value:
                    header_lower = str(header_value).lower()
                    if 'id' in header_lower:
                        headers['id'] = col
                    elif any(x in header_lower for x in ['requirement', 'description', 'text']):
                        headers['text'] = col
                    elif 'priority' in header_lower:
                        headers['priority'] = col
                    elif 'category' in header_lower:
                        headers['category'] = col
            
            # Process data rows
            for row_num in range(2, worksheet.max_row + 1):
                req_id = str(worksheet.cell(row_num, headers.get('id', 1)).value or f"REQ{row_num-1:03d}")
                req_text = str(worksheet.cell(row_num, headers.get('text', 2)).value or "")
                priority = str(worksheet.cell(row_num, headers.get('priority', 3)).value or "Medium")
                category = str(worksheet.cell(row_num, headers.get('category', 4)).value or "Functional")
                
                if req_text.strip() and req_text != "None":
                    requirement = Requirement(
                        id=req_id,
                        text=req_text.strip(),
                        priority=priority,
                        category=category,
                        source=RequirementSource.EXCEL_FILE,
                        line_number=row_num
                    )
                    requirements.append(requirement)
                    
        except ImportError:
            print("openpyxl not installed. Install with: pip install openpyxl")
        except Exception as e:
            print(f"Error processing Excel file: {e}")
        
        return requirements
    
    @staticmethod
    def _process_json_file(file_path: str) -> List[Requirement]:
        """Process JSON file containing requirements"""
        requirements = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            
            # Handle different JSON structures
            if isinstance(data, list):
                # Array of requirement objects
                for i, item in enumerate(data):
                    if isinstance(item, dict):
                        req_id = item.get('id', f"REQ{i+1:03d}")
                        req_text = item.get('text', item.get('requirement', item.get('description', "")))
                        priority = item.get('priority', "Medium")
                        category = item.get('category', "Functional")
                        
                        if req_text:
                            requirement = Requirement(
                                id=req_id,
                                text=req_text,
                                priority=priority,
                                category=category,
                                source=RequirementSource.JSON_FILE,
                                line_number=i+1
                            )
                            requirements.append(requirement)
            
            elif isinstance(data, dict):
                # Object with requirements key or direct requirement objects
                reqs_data = data.get('requirements', data)
                
                if isinstance(reqs_data, list):
                    for i, item in enumerate(reqs_data):
                        req_id = item.get('id', f"REQ{i+1:03d}")
                        req_text = item.get('text', item.get('requirement', ""))
                        priority = item.get('priority', "Medium")
                        category = item.get('category', "Functional")
                        
                        if req_text:
                            requirement = Requirement(
                                id=req_id,
                                text=req_text,
                                priority=priority,
                                category=category,
                                source=RequirementSource.JSON_FILE,
                                line_number=i+1
                            )
                            requirements.append(requirement)
                elif isinstance(reqs_data, dict):
                    # Requirements as key-value pairs
                    for i, (key, value) in enumerate(reqs_data.items()):
                        req_text = value if isinstance(value, str) else value.get('text', str(value))
                        requirement = Requirement(
                            id=key,
                            text=req_text,
                            priority="Medium",
                            category="Functional",
                            source=RequirementSource.JSON_FILE,
                            line_number=i+1
                        )
                        requirements.append(requirement)
                        
        except Exception as e:
            print(f"Error processing JSON file: {e}")
        
        return requirements
    
    @staticmethod
    def _process_text_file(file_path: str) -> List[Requirement]:
        """Process text file containing requirements"""
        requirements = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            
            lines = content.split('\n')
            req_counter = 1
            
            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue
                
                # Check if line starts with requirement identifier
                req_match = re.match(r'^(REQ[_-]?\d+|R\d+|\d+\.|\w+\d+):?\s*(.+)', line, re.IGNORECASE)
                if req_match:
                    req_id = req_match.group(1).replace(':', '')
                    req_text = req_match.group(2)
                    
                    requirement = Requirement(
                        id=req_id,
                        text=req_text,
                        priority="Medium",
                        category="Functional",
                        source=RequirementSource.TEXT_FILE,
                        line_number=line_num
                    )
                    requirements.append(requirement)
                    
                elif line.startswith(('-', '*', '•')):
                    # Bullet point format
                    req_text = line[1:].strip()
                    req_id = f"REQ{req_counter:03d}"
                    req_counter += 1
                    
                    requirement = Requirement(
                        id=req_id,
                        text=req_text,
                        priority="Medium",
                        category="Functional",
                        source=RequirementSource.TEXT_FILE,
                        line_number=line_num
                    )
                    requirements.append(requirement)
                    
                elif len(line) > 20:  # Assume longer lines are requirements
                    req_id = f"REQ{req_counter:03d}"
                    req_counter += 1
                    
                    requirement = Requirement(
                        id=req_id,
                        text=line,
                        priority="Medium",
                        category="Functional",
                        source=RequirementSource.TEXT_FILE,
                        line_number=line_num
                    )
                    requirements.append(requirement)
                    
        except Exception as e:
            print(f"Error processing text file: {e}")
        
        return requirements
    
    @staticmethod
    def _process_word_file(file_path: str) -> List[Requirement]:
        """Process Word document containing requirements"""
        requirements = []
        
        try:
            import docx
            
            doc = docx.Document(file_path)
            req_counter = 1
            
            for para_num, paragraph in enumerate(doc.paragraphs, 1):
                text = paragraph.text.strip()
                if not text or len(text) < 10:
                    continue
                
                # Check if paragraph starts with requirement identifier
                req_match = re.match(r'^(REQ[_-]?\d+|R\d+|\d+\.|\w+\d+):?\s*(.+)', text, re.IGNORECASE)
                if req_match:
                    req_id = req_match.group(1).replace(':', '')
                    req_text = req_match.group(2)
                else:
                    req_id = f"REQ{req_counter:03d}"
                    req_text = text
                    req_counter += 1
                
                # Check if this looks like a requirement (contains "shall", "must", "will", etc.)
                if any(keyword in text.lower() for keyword in ['shall', 'must', 'will', 'should', 'require']):
                    requirement = Requirement(
                        id=req_id,
                        text=req_text,
                        priority="Medium",
                        category="Functional",
                        source=RequirementSource.WORD_FILE,
                        line_number=para_num
                    )
                    requirements.append(requirement)
                    
        except ImportError:
            print("python-docx not installed. Install with: pip install python-docx")
        except Exception as e:
            print(f"Error processing Word file: {e}")
        
        return requirements

class AgentState(TypedDict):
    """State for the LangGraph agent"""
    messages: Annotated[List[BaseMessage], operator.add]
    requirements: List[Requirement]
    parsed_requirements: List[Dict[str, Any]]
    test_cases: List[TestCase]
    current_requirement: Optional[Dict[str, Any]]
    analysis_complete: bool
    generation_complete: bool
    validation_complete: bool
    next_action: str
    file_sources: List[str]

class AgenticTestCaseGenerator:
    """Agentic test case generator using LangGraph"""
    
    #def __init__(self, llm_model: str = "gpt-4o", api_key: Optional[str] = None):
        # For demo, we'll skip LLM initialization
    def __init__(self, llm_model: str = "gpt-4o", api_key: Optional[str] = None):
        self.llm = ChatOpenAI(
            model=llm_model,
            api_key=OPENAI_API_KEY,
            temperature=0.1
        )
        self.graph = self._build_graph()    
        #pass
        
    def _build_graph(self):
        """Build the LangGraph workflow"""
        workflow = StateGraph(AgentState)
        
        # Add nodes
        workflow.add_node("analyze_requirements", self._analyze_requirements)
        workflow.add_node("generate_test_cases", self._generate_test_cases)
        workflow.add_node("validate_test_cases", self._validate_test_cases)
        workflow.add_node("optimize_test_suite", self._optimize_test_suite)
        workflow.add_node("finalize_output", self._finalize_output)
        
        # Define edges and conditions
        workflow.add_edge("analyze_requirements", "generate_test_cases")
        workflow.add_edge("generate_test_cases", "validate_test_cases")
        workflow.add_edge("validate_test_cases", "optimize_test_suite")
        workflow.add_edge("optimize_test_suite", "finalize_output")
        workflow.add_edge("finalize_output", END)
        
        # Set entry point
        workflow.set_entry_point("analyze_requirements")
        
        return workflow.compile()
    
    def _analyze_requirements(self, state: AgentState) -> AgentState:
        """Analyze requirements"""
        print("🔍 Analyzing requirements...")
        
        parsed_requirements = []
        
        for req in state["requirements"]:
            analysis = self._basic_requirement_analysis(req)
            parsed_requirements.append(analysis)
        
        state["parsed_requirements"] = parsed_requirements
        state["analysis_complete"] = True
        state["messages"].append(AIMessage(content=f"Analyzed {len(parsed_requirements)} requirements"))
        
        return state
    
    def _generate_test_cases(self, state: AgentState) -> AgentState:
        """Generate test cases"""
        print("🚀 Generating test cases...")
        
        all_test_cases = []
        
        for req_analysis in state["parsed_requirements"]:
            print(f"   Generating test cases for {req_analysis['id']}...")
            test_cases = self._generate_enhanced_template_test_cases(req_analysis)
            
            # Set requirement_id for traceability
            for tc in test_cases:
                tc.requirement_id = req_analysis['id']
            
            print(f"   ✅ Generated {len(test_cases)} test cases for {req_analysis['id']}")
            all_test_cases.extend(test_cases)
        
        state["test_cases"] = all_test_cases
        state["generation_complete"] = True
        state["messages"].append(AIMessage(content=f"Generated {len(all_test_cases)} test cases"))
        
        return state
    
    def _validate_test_cases(self, state: AgentState) -> AgentState:
        """Validate and improve test cases"""
        print("✅ Validating test cases...")
        
        validated_cases = []
        
        for test_case in state["test_cases"]:
            validation_issues = self._validate_single_test_case(test_case)
            
            if validation_issues:
                fixed_case = self._fix_test_case_issues(test_case, validation_issues)
                validated_cases.append(fixed_case)
            else:
                validated_cases.append(test_case)
        
        state["test_cases"] = validated_cases
        state["validation_complete"] = True
        state["messages"].append(AIMessage(content=f"Validated {len(validated_cases)} test cases"))
        
        return state
    
    def _optimize_test_suite(self, state: AgentState) -> AgentState:
        """Optimize the test suite for coverage and efficiency"""
        print("⚡ Optimizing test suite...")
        
        # Remove duplicates
        unique_cases = self._remove_duplicate_test_cases(state["test_cases"])
        
        # Prioritize test cases
        prioritized_cases = self._prioritize_test_cases(unique_cases)
        
        # Add traceability
        traceable_cases = self._add_traceability(prioritized_cases, state["parsed_requirements"])
        
        state["test_cases"] = traceable_cases
        state["messages"].append(AIMessage(content=f"Optimized test suite: {len(traceable_cases)} test cases"))
        
        return state
    
    def _finalize_output(self, state: AgentState) -> AgentState:
        """Finalize the output"""
        print("🎯 Finalizing output...")
        
        # Generate summary statistics
        summary = self._generate_test_suite_summary(state["test_cases"])
        
        state["messages"].append(AIMessage(content=f"Test suite complete: {summary}"))
        
        return state
    
    def _basic_requirement_analysis(self, req: Requirement) -> Dict[str, Any]:
        """Basic analysis when LLM is not available"""
        return {
            "id": req.id,
            "original_text": req.text,
            "priority": req.priority,
            "category": req.category,
            "source": req.source.value,
            "line_number": req.line_number,
            "functional_aspects": ["Core functionality"],
            "non_functional_aspects": [],
            "business_rules": [],
            "error_conditions": ["Invalid input"],
            "risk_level": "Medium",
            "testability": "High",
            "dependencies": []
        }
    
    def _generate_enhanced_template_test_cases(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate enhanced template test cases with requirement-specific logic"""
        test_cases = []
        req_id = req_analysis["id"]
        req_text = req_analysis["original_text"].lower()
        
        # Detect requirement type and generate appropriate test cases
        if "password" in req_text:
            test_cases.extend(self._generate_password_test_cases(req_analysis))
        elif "email" in req_text:
            test_cases.extend(self._generate_email_test_cases(req_analysis))
        elif "login" in req_text or "authentication" in req_text:
            test_cases.extend(self._generate_login_test_cases(req_analysis))
        elif "rate limit" in req_text or "brute force" in req_text:
            test_cases.extend(self._generate_security_test_cases(req_analysis))
        else:
            # Generic test cases
            test_cases.extend(self._generate_generic_test_cases(req_analysis))
        
        return test_cases
    
    def _generate_password_test_cases(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate specific test cases for password requirements"""
        req_id = req_analysis["id"]
        test_cases = []
        
        # Positive test case - valid password
        positive_case = TestCase(
            id=f"{req_id}_POSITIVE_001",
            title=f"Verify {req_id} - Valid password meets all criteria",
            description="Test that a password meeting all requirements is accepted",
            test_type=TestType.POSITIVE,
            preconditions=["User registration/password change form is accessible"],
            test_steps=[
                "Navigate to password input field",
                "Enter password: 'TestPass123!'",
                "Submit the form",
                "Verify password is accepted"
            ],
            expected_result="Password is accepted and validation passes",
            priority="High",
            tags=["password", "validation", "positive"],
            risk_level="High"
        )
        test_cases.append(positive_case)
        
        # Negative test cases
        negative_cases = [
            {
                "suffix": "TOO_SHORT",
                "title": "Password too short (7 characters)",
                "password": "Test1!",
                "expected": "Password rejected - too short"
            },
            {
                "suffix": "NO_UPPERCASE",
                "title": "Missing uppercase letter",
                "password": "testpass123!",
                "expected": "Password rejected - missing uppercase"
            },
            {
                "suffix": "NO_DIGIT",
                "title": "Missing digit",
                "password": "TestPassword!",
                "expected": "Password rejected - missing digit"
            }
        ]
        
        for i, case_data in enumerate(negative_cases, 1):
            negative_case = TestCase(
                id=f"{req_id}_NEGATIVE_{i:03d}",
                title=f"Verify {req_id} - {case_data['title']}",
                description=f"Test that password validation rejects: {case_data['title'].lower()}",
                test_type=TestType.NEGATIVE,
                preconditions=["User registration/password change form is accessible"],
                test_steps=[
                    "Navigate to password input field",
                    f"Enter invalid password: '{case_data['password']}'",
                    "Submit the form",
                    "Verify appropriate error message is displayed"
                ],
                expected_result=case_data["expected"],
                priority="High",
                tags=["password", "validation", "negative"],
                risk_level="Medium"
            )
            test_cases.append(negative_case)
        
        return test_cases
    
    def _generate_email_test_cases(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate test cases for email validation requirements"""
        req_id = req_analysis["id"]
        test_cases = []
        
        positive_case = TestCase(
            id=f"{req_id}_POSITIVE_001",
            title=f"Verify {req_id} - Valid email format accepted",
            description="Test that valid email addresses are accepted",
            test_type=TestType.POSITIVE,
            preconditions=["Email input field is accessible"],
            test_steps=[
                "Navigate to email input field",
                "Enter valid email: 'user@example.com'",
                "Submit the form",
                "Verify email is accepted"
            ],
            expected_result="Valid email is accepted without errors",
            priority="High",
            tags=["email", "validation", "positive"]
        )
        test_cases.append(positive_case)
        
        return test_cases
    
    def _generate_login_test_cases(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate test cases for login requirements"""
        req_id = req_analysis["id"]
        test_cases = []
        
        positive_case = TestCase(
            id=f"{req_id}_POSITIVE_001",
            title=f"Verify {req_id} - Successful login with valid credentials",
            description="Test successful login with correct username and password",
            test_type=TestType.POSITIVE,
            preconditions=["Valid user account exists", "Login page is accessible"],
            test_steps=[
                "Navigate to login page",
                "Enter valid username",
                "Enter valid password", 
                "Click login button",
                "Verify successful login"
            ],
            expected_result="User is successfully logged in and redirected to dashboard",
            priority="High",
            tags=["login", "authentication", "positive"]
        )
        test_cases.append(positive_case)
        
        return test_cases
    
    def _generate_security_test_cases(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate test cases for security requirements"""
        req_id = req_analysis["id"]
        test_cases = []
        
        security_case = TestCase(
            id=f"{req_id}_SECURITY_001",
            title=f"Verify {req_id} - Rate limiting protection",
            description="Test that rate limiting prevents brute force attacks",
            test_type=TestType.SECURITY,
            preconditions=["Application is running", "Test user account exists"],
            test_steps=[
                "Attempt multiple rapid requests/login attempts",
                "Exceed the defined rate limit",
                "Verify rate limiting is enforced",
                "Check appropriate error response"
            ],
            expected_result="Rate limiting blocks excessive requests and returns appropriate error",
            priority="High",
            tags=["security", "rate_limiting", "brute_force"]
        )
        test_cases.append(security_case)
        
        return test_cases
    
    def _generate_generic_test_cases(self, req_analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate generic test cases for any requirement"""
        req_id = req_analysis["id"]
        test_cases = []
        
        # Generic positive case
        positive_case = TestCase(
            id=f"{req_id}_POSITIVE_001",
            title=f"Verify {req_id} - Valid scenario",
            description=f"Test valid scenario for requirement {req_id}",
            test_type=TestType.POSITIVE,
            preconditions=["System is ready", "Valid test data available"],
            test_steps=[
                "Prepare valid test data",
                "Execute the required functionality",
                "Verify expected behavior"
            ],
            expected_result="System behaves as specified in the requirement",
            priority="High",
            tags=["functional", "positive"]
        )
        test_cases.append(positive_case)
        
        # Generic negative case
        negative_case = TestCase(
            id=f"{req_id}_NEGATIVE_001", 
            title=f"Verify {req_id} - Invalid scenario",
            description=f"Test invalid scenario for requirement {req_id}",
            test_type=TestType.NEGATIVE,
            preconditions=["System is ready"],
            test_steps=[
                "Prepare invalid test data",
                "Attempt to execute functionality",
                "Verify error handling"
            ],
            expected_result="System handles error appropriately with proper error messages",
            priority="Medium",
            tags=["functional", "negative"]
        )
        test_cases.append(negative_case)
        
        return test_cases
    
    def _validate_single_test_case(self, test_case: TestCase) -> List[str]:
        """Validate a single test case and return issues"""
        issues = []
        
        if not test_case.test_steps:
            issues.append("Missing test steps")
        
        if not test_case.expected_result:
            issues.append("Missing expected result")
        
        if len(test_case.title) < 10:
            issues.append("Title too short")
        
        return issues
    
    def _fix_test_case_issues(self, test_case: TestCase, issues: List[str]) -> TestCase:
        """Fix identified issues in test case"""
        if "Missing test steps" in issues:
            test_case.test_steps = ["Execute test scenario", "Verify results"]
        
        if "Missing expected result" in issues:
            test_case.expected_result = "System behaves as expected"
        
        if "Title too short" in issues:
            test_case.title = f"Test Case: {test_case.title}"
        
        return test_case
    
    def _remove_duplicate_test_cases(self, test_cases: List[TestCase]) -> List[TestCase]:
        """Remove duplicate test cases"""
        seen_titles = set()
        unique_cases = []
        
        for case in test_cases:
            if case.title not in seen_titles:
                seen_titles.add(case.title)
                unique_cases.append(case)
        
        return unique_cases
    
    def _prioritize_test_cases(self, test_cases: List[TestCase]) -> List[TestCase]:
        """Prioritize test cases based on risk and type"""
        priority_order = {"High": 1, "Medium": 2, "Low": 3}
        type_order = {
            TestType.POSITIVE: 1,
            TestType.SECURITY: 2,
            TestType.NEGATIVE: 3,
            TestType.BOUNDARY: 4,
            TestType.EDGE_CASE: 5,
            TestType.INTEGRATION: 6
        }
        
        return sorted(test_cases, key=lambda x: (
            priority_order.get(x.priority, 2),
            type_order.get(x.test_type, 5)
        ))
    
    def _add_traceability(self, test_cases: List[TestCase], requirements: List[Dict[str, Any]]) -> List[TestCase]:
        """Add traceability links between test cases and requirements"""
        req_map = {req["id"]: req for req in requirements}
        
        for case in test_cases:
            req_id = case.id.split("_")[0]  # Extract requirement ID from test case ID
            if req_id in req_map:
                case.tags.append(f"traces_to_{req_id}")
        
        return test_cases
    
    def _generate_test_suite_summary(self, test_cases: List[TestCase]) -> str:
        """Generate summary of the test suite"""
        total = len(test_cases)
        by_type = {}
        by_priority = {}
        
        for case in test_cases:
            by_type[case.test_type.value] = by_type.get(case.test_type.value, 0) + 1
            by_priority[case.priority] = by_priority.get(case.priority, 0) + 1
        
        return f"{total} total test cases. Types: {by_type}. Priorities: {by_priority}"
    
    def generate_test_cases_from_files(self, file_paths: List[str]) -> List[TestCase]:
        """Generate test cases from uploaded requirement files"""
        all_requirements = []
        
        print(f"📁 Processing {len(file_paths)} requirement files...")
        
        for file_path in file_paths:
            print(f"   📄 Processing {file_path}...")
            requirements = FileProcessor.process_file(file_path)
            
            if requirements:
                print(f"   ✅ Found {len(requirements)} requirements in {file_path}")
                all_requirements.extend(requirements)
            else:
                print(f"   ⚠️ No requirements found in {file_path}")
        
        if not all_requirements:
            print("❌ No requirements found in any files")
            return []
        
        print(f"📋 Total requirements loaded: {len(all_requirements)}")
        
        # Convert to expected format and generate test cases
        return self._generate_test_cases_from_requirements(all_requirements, file_paths)
    
    def generate_test_cases_from_manual_input(self, requirements: List[tuple]) -> List[TestCase]:
        """Generate test cases from manually entered requirements"""
        # Convert tuples to Requirement objects
        req_objects = []
        for req_id, req_text in requirements:
            req_obj = Requirement(
                id=req_id,
                text=req_text,
                source=RequirementSource.MANUAL
            )
            req_objects.append(req_obj)
        
        return self._generate_test_cases_from_requirements(req_objects, [])
    
    def _generate_test_cases_from_requirements(self, requirements: List[Requirement], file_sources: List[str]) -> List[TestCase]:
        """Internal method to generate test cases from Requirement objects"""
        # Build the graph
        self.graph = self._build_graph()
        
        # Initial state
        initial_state = {
            "messages": [HumanMessage(content="Starting test case generation from file uploads")],
            "requirements": requirements,
            "parsed_requirements": [],
            "test_cases": [],
            "current_requirement": None,
            "analysis_complete": False,
            "generation_complete": False,
            "validation_complete": False,
            "next_action": "analyze",
            "file_sources": file_sources
        }
        
        # Run the agent workflow
        final_state = self.graph.invoke(initial_state)
        
        return final_state["test_cases"]
    
    def export_to_json(self, test_cases: List[TestCase], filename: str = "test_cases.json"):
        """Export test cases to JSON"""
        cases_dict = [asdict(case) for case in test_cases]
        # Convert enums to strings
        for case in cases_dict:
            case["test_type"] = case["test_type"].value if hasattr(case["test_type"], 'value') else case["test_type"]
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(cases_dict, f, indent=2, ensure_ascii=False)
    
    def export_to_csv(self, test_cases: List[TestCase], filename: str = "test_cases.csv"):
        """Export test cases to CSV format"""
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Test ID', 'Title', 'Description', 'Type', 'Priority', 
                         'Preconditions', 'Test Steps', 'Expected Result', 'Tags',
                         'Risk Level', 'Requirement ID', 'Automation Feasible']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for tc in test_cases:
                writer.writerow({
                    'Test ID': tc.id,
                    'Title': tc.title,
                    'Description': tc.description,
                    'Type': tc.test_type.value,
                    'Priority': tc.priority,
                    'Preconditions': '; '.join(tc.preconditions),
                    'Test Steps': '; '.join(tc.test_steps),
                    'Expected Result': tc.expected_result,
                    'Tags': ', '.join(tc.tags),
                    'Risk Level': tc.risk_level,
                    'Requirement ID': tc.requirement_id,
                    'Automation Feasible': tc.automation_feasible
                })
    
    def export_to_excel(self, test_cases: List[TestCase], filename: str = "test_cases.xlsx"):
        """Export test cases to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Cases"
            
            # Headers
            headers = ['Test ID', 'Title', 'Description', 'Type', 'Priority', 
                      'Preconditions', 'Test Steps', 'Expected Result', 'Tags',
                      'Risk Level', 'Requirement ID', 'Automation Feasible']
            
            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")
            
            # Add data
            for row, tc in enumerate(test_cases, 2):
                ws.cell(row=row, column=1, value=tc.id)
                ws.cell(row=row, column=2, value=tc.title)
                ws.cell(row=row, column=3, value=tc.description)
                ws.cell(row=row, column=4, value=tc.test_type.value)
                ws.cell(row=row, column=5, value=tc.priority)
                ws.cell(row=row, column=6, value='; '.join(tc.preconditions))
                ws.cell(row=row, column=7, value='; '.join(tc.test_steps))
                ws.cell(row=row, column=8, value=tc.expected_result)
                ws.cell(row=row, column=9, value=', '.join(tc.tags))
                ws.cell(row=row, column=10, value=tc.risk_level)
                ws.cell(row=row, column=11, value=tc.requirement_id)
                ws.cell(row=row, column=12, value=tc.automation_feasible)
                
                # Color code by priority
                if tc.priority == "High":
                    fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                elif tc.priority == "Medium":
                    fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                else:  # Low
                    fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
                
                for col in range(1, len(headers) + 1):
                    ws.cell(row=row, column=col).fill = fill
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save workbook
            wb.save(filename)
            
        except ImportError:
            print("openpyxl not installed. Install with: pip install openpyxl")
            # Fallback to CSV
            self.export_to_csv(test_cases, filename.replace('.xlsx', '.csv'))
        except Exception as e:
            print(f"Error exporting to Excel: {e}")
            # Fallback to CSV
            self.export_to_csv(test_cases, filename.replace('.xlsx', '.csv'))

    def create_file_upload_interface(self):
        """Create a simple file upload interface description"""
        interface_info = {
            "supported_formats": {
                "CSV": {
                    "extensions": [".csv"],
                    "expected_columns": ["ID", "Requirement", "Priority", "Category"],
                    "example": "REQ001,User shall login with valid credentials,High,Authentication"
                },
                "Excel": {
                    "extensions": [".xlsx", ".xls"],
                    "expected_columns": ["ID", "Requirement", "Priority", "Category"],
                    "note": "Uses first worksheet, expects headers in first row"
                },
                "JSON": {
                    "extensions": [".json"],
                    "expected_structure": {
                        "requirements": [
                            {
                                "id": "REQ001",
                                "text": "Requirement description",
                                "priority": "High",
                                "category": "Functional"
                            }
                        ]
                    }
                },
                "Text": {
                    "extensions": [".txt"],
                    "formats": [
                        "REQ001: Requirement description",
                        "- Requirement as bullet point",
                        "Plain requirement text (auto-numbered)"
                    ]
                },
                "Word": {
                    "extensions": [".docx", ".doc"],
                    "note": "Extracts paragraphs containing requirement keywords (shall, must, will, etc.)"
                }
            },
            "usage_examples": {
                "single_file": "generator.generate_test_cases_from_files(['requirements.csv'])",
                "multiple_files": "generator.generate_test_cases_from_files(['req1.csv', 'req2.xlsx', 'req3.json'])",
                "mixed_sources": "Combine file uploads with manual requirements"
            }
        }
        
        return interface_info
    
    def print_test_cases(self, test_cases: List[TestCase]):
        """Print test cases in a readable format"""
        for tc in test_cases:
            print(f"\n{'='*60}")
            print(f"Test ID: {tc.id}")
            print(f"Title: {tc.title}")
            print(f"Type: {tc.test_type.value.upper()}")
            print(f"Priority: {tc.priority}")
            print(f"Risk Level: {tc.risk_level}")
            print(f"Description: {tc.description}")
            print(f"\nPreconditions:")
            for i, precond in enumerate(tc.preconditions, 1):
                print(f"  {i}. {precond}")
            print(f"\nTest Steps:")
            for i, step in enumerate(tc.test_steps, 1):
                print(f"  {i}. {step}")
            print(f"\nExpected Result: {tc.expected_result}")
            print(f"Tags: {', '.join(tc.tags)}")
            print(f"Automation Feasible: {tc.automation_feasible}")


# Demo/Main function
def main():
    """Example usage of the Enhanced Agentic Test Case Generator with File Upload Support"""
    
    def create_sample_files():
        """Create sample requirement files for demonstration"""
        
        # Sample CSV file
        csv_content = """ID,Requirement,Priority,Category
REQ001,"The system shall validate user email addresses using RFC 5322 standard and reject invalid formats with appropriate error messages",High,Authentication
REQ002,"User passwords must be between 8 and 128 characters long and contain at least one uppercase letter, one lowercase letter, one digit, and one special character",High,Security
REQ003,"The system shall implement rate limiting to prevent brute force attacks, allowing maximum 5 login attempts per IP address within 15 minutes",Medium,Security
REQ004,"When a user enters invalid credentials 3 consecutive times, the account shall be temporarily locked for 30 minutes",High,Security
REQ005,"The system shall log all authentication events including successful logins, failed attempts, and account lockouts for security audit purposes",Medium,Logging"""
        
        with open('sample_requirements.csv', 'w', encoding='utf-8') as f:
            f.write(csv_content)
        
        # Sample JSON file
        json_content = {
            "requirements": [
                {
                    "id": "REQ006",
                    "text": "The system shall provide a password reset functionality via email verification",
                    "priority": "Medium",
                    "category": "Authentication"
                },
                {
                    "id": "REQ007", 
                    "text": "Users shall be able to enable two-factor authentication using TOTP",
                    "priority": "High",
                    "category": "Security"
                }
            ]
        }
        
        with open('sample_requirements.json', 'w', encoding='utf-8') as f:
            json.dump(json_content, f, indent=2)
        
        # Sample text file
        text_content = """REQ008: The system shall maintain user session for maximum 8 hours of inactivity
REQ009: All user inputs shall be sanitized to prevent SQL injection attacks
- The system must encrypt all sensitive data at rest using AES-256
- User profile data shall be editable by authenticated users only
REQ012: The application shall support concurrent users up to 10,000 simultaneous sessions"""
        
        with open('sample_requirements.txt', 'w', encoding='utf-8') as f:
            f.write(text_content)
        
        return ['sample_requirements.csv', 'sample_requirements.json', 'sample_requirements.txt']
    
    print("🤖 Starting Enhanced Agentic Test Case Generation with File Upload Support!")
    print("📝 Note: This demo uses template-based generation. For full AI capabilities, provide API keys.")
    print("=" * 80)
    
    generator = AgenticTestCaseGenerator()
    
    # Show supported file formats
    print("📁 Supported File Formats:")
    interface_info = generator.create_file_upload_interface()
    for format_name, format_info in interface_info["supported_formats"].items():
        print(f"   📄 {format_name}: {format_info['extensions']}")
    
    print("\n🔧 Creating sample requirement files...")
    sample_files = create_sample_files()
    print(f"   ✅ Created: {', '.join(sample_files)}")
    
    try:
        print(f"\n📂 Processing requirement files...")
        # Generate test cases from files
        test_cases = generator.generate_test_cases_from_files(sample_files)
        
        # Display results
        print(f"\n🎉 Generated {len(test_cases)} test cases from file uploads!")
        
        # Summary by source
        print(f"\n📊 Test cases by source:")
        by_source = {}
        for tc in test_cases:
            source = "Unknown"
            if any(req_id in tc.id for req_id in ["REQ001", "REQ002", "REQ003", "REQ004", "REQ005"]):
                source = "CSV File"
            elif any(req_id in tc.id for req_id in ["REQ006", "REQ007"]):
                source = "JSON File"
            elif any(req_id in tc.id for req_id in ["REQ008", "REQ009", "REQ010", "REQ011", "REQ012"]):
                source = "Text File"
            
            by_source[source] = by_source.get(source, 0) + 1
        
        for source, count in by_source.items():
            print(f"   {source}: {count} test cases")
        
        # Summary by requirement
        print(f"\n📋 Test cases by requirement:")
        by_req = {}
        for tc in test_cases:
            req_id = tc.requirement_id or tc.id.split("_")[0]
            by_req[req_id] = by_req.get(req_id, 0) + 1
        
        for req_id, count in sorted(by_req.items()):
            print(f"   {req_id}: {count} test cases")
        
        # Show detailed test cases for password requirement
        print(f"\n🔐 Detailed view - Password requirement (REQ002) test cases:")
        req002_cases = [tc for tc in test_cases if tc.requirement_id == "REQ002" or tc.id.startswith("REQ002")]
        for tc in req002_cases:
            print(f"   - {tc.id}: {tc.title}")
            print(f"     Type: {tc.test_type.value}, Priority: {tc.priority}")
        
        # Export results in different formats
        print(f"\n💾 Exporting test cases...")
        generator.export_to_json(test_cases, "file_based_test_cases.json")
        print(f"   ✅ Exported to JSON: file_based_test_cases.json")
        
        generator.export_to_csv(test_cases, "file_based_test_cases.csv")
        print(f"   ✅ Exported to CSV: file_based_test_cases.csv")
        
        try:
            generator.export_to_excel(test_cases, "file_based_test_cases.xlsx")
            print(f"   ✅ Exported to Excel: file_based_test_cases.xlsx")
        except Exception as e:
            print(f"   ⚠️ Excel export failed: {e}")
        
        # Show file upload usage examples
        print(f"\n📚 Usage Examples:")
        #print(f"   Single file: generator.generate_test_cases_from_files(['requirements.csv'])")
        print(f"   Single file: generator.generate_test_cases_from_files(['BRD.csv'])")
        print(f"   Multiple files: generator.generate_test_cases_from_files(['req1.csv', 'req2.xlsx'])")
        print(f"   Manual input: generator.generate_test_cases_from_manual_input([('REQ001', 'text')])")
        
        # Clean up sample files
        import os
        for file in sample_files:
            try:
                os.remove(file)
            except:
                pass
        print(f"\n🧹 Cleaned up sample files")
        
    except Exception as e:
        print(f"❌ Error during test case generation: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()