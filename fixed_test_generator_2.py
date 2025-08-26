import re
import json
import csv
import io
from typing import List, Dict, Any, Optional, Tuple, Set
from dataclasses import dataclass, asdict
from enum import Enum
import os
from datetime import datetime

class TestType(Enum):
    POSITIVE = "positive"
    NEGATIVE = "negative"
    BOUNDARY = "boundary"
    EDGE_CASE = "edge_case"
    INTEGRATION = "integration"
    SECURITY = "security"
    PERFORMANCE = "performance"
    USABILITY = "usability"

class RequirementSource(Enum):
    MANUAL = "manual"
    CSV_FILE = "csv_file"
    EXCEL_FILE = "excel_file"
    JSON_FILE = "json_file"
    TEXT_FILE = "text_file"
    WORD_FILE = "word_file"

class RequirementType(Enum):
    FUNCTIONAL = "functional"
    NON_FUNCTIONAL = "non_functional"
    SECURITY = "security"
    PERFORMANCE = "performance"
    USABILITY = "usability"
    BUSINESS = "business"
    TECHNICAL = "technical"

@dataclass
class Requirement:
    """Represents a single requirement"""
    id: str
    text: str
    priority: str = "Medium"
    category: str = "Functional"
    source: RequirementSource = RequirementSource.MANUAL
    line_number: Optional[int] = None
    requirement_type: RequirementType = RequirementType.FUNCTIONAL
    complexity: str = "Medium"
    
@dataclass
class TestCase:
    """Represents a single test case"""
    id: str
    title: str
    description: str
    test_type: TestType
    preconditions: List[str]
    test_steps: List[str]
    expected_result: str
    priority: str = "Medium"
    tags: List[str] = None
    risk_level: str = "Medium"
    automation_feasible: bool = True
    requirement_id: str = ""
    estimated_time: int = 30  # minutes
    test_data: List[str] = None
    
    def __post_init__(self):
        if self.tags is None:
            self.tags = []
        if self.test_data is None:
            self.test_data = []

class RequirementAnalyzer:
    """Analyzes requirements to determine type, complexity, and testing approach"""
    
    def __init__(self):
        self.security_keywords = [
            'password', 'encrypt', 'authentication', 'authorization', 'security',
            'access control', 'permission', 'role', 'token', 'session', 'login',
            'logout', 'brute force', 'rate limit', 'ssl', 'tls', 'certificate',
            'vulnerability', 'attack', 'malicious', 'injection', 'xss', 'csrf'
        ]
        
        self.performance_keywords = [
            'performance', 'speed', 'response time', 'latency', 'throughput',
            'concurrent', 'load', 'stress', 'scalability', 'memory', 'cpu',
            'database', 'query', 'cache', 'optimization', 'benchmark'
        ]
        
        self.usability_keywords = [
            'user interface', 'ui', 'ux', 'usability', 'accessibility',
            'user experience', 'navigation', 'layout', 'design', 'intuitive',
            'user-friendly', 'responsive', 'mobile', 'browser', 'display'
        ]
        
        self.business_keywords = [
            'business rule', 'workflow', 'process', 'approval', 'notification',
            'report', 'dashboard', 'integration', 'api', 'third party',
            'compliance', 'regulation', 'audit', 'logging'
        ]
        
        self.data_keywords = [
            'data', 'database', 'storage', 'backup', 'recovery', 'migration',
            'import', 'export', 'sync', 'replication', 'validation',
            'format', 'field', 'record', 'table', 'query'
        ]
    
    def analyze_requirement(self, requirement: Requirement) -> Dict[str, Any]:
        """Analyze a requirement and return detailed analysis"""
        text_lower = requirement.text.lower()
        
        analysis = {
            'id': requirement.id,
            'original_text': requirement.text,
            'priority': requirement.priority,
            'category': requirement.category,
            'source': requirement.source.value,
            'line_number': requirement.line_number,
            'requirement_type': self._determine_requirement_type(text_lower),
            'complexity': self._assess_complexity(text_lower),
            'testing_approaches': self._suggest_testing_approaches(text_lower),
            'risk_factors': self._identify_risk_factors(text_lower),
            'test_data_needs': self._identify_test_data_needs(text_lower),
            'integration_points': self._find_integration_points(text_lower),
            'performance_considerations': self._assess_performance_needs(text_lower),
            'security_considerations': self._assess_security_needs(text_lower),
            'boundary_conditions': self._identify_boundary_conditions(text_lower),
            'error_scenarios': self._identify_error_scenarios(text_lower)
        }
        
        return analysis
    
    def _determine_requirement_type(self, text: str) -> str:
        """Determine the type of requirement"""
        if any(keyword in text for keyword in self.security_keywords):
            return RequirementType.SECURITY.value
        elif any(keyword in text for keyword in self.performance_keywords):
            return RequirementType.PERFORMANCE.value
        elif any(keyword in text for keyword in self.usability_keywords):
            return RequirementType.USABILITY.value
        elif any(keyword in text for keyword in self.business_keywords):
            return RequirementType.BUSINESS.value
        elif any(keyword in text for keyword in self.data_keywords):
            return RequirementType.TECHNICAL.value
        else:
            return RequirementType.FUNCTIONAL.value
    
    def _assess_complexity(self, text: str) -> str:
        """Assess the complexity of the requirement"""
        complexity_indicators = {
            'high': ['complex', 'multiple', 'integration', 'workflow', 'algorithm', 
                    'calculation', 'real-time', 'concurrent', 'distributed'],
            'low': ['simple', 'basic', 'single', 'display', 'show', 'list']
        }
        
        if any(indicator in text for indicator in complexity_indicators['high']):
            return "High"
        elif any(indicator in text for indicator in complexity_indicators['low']):
            return "Low"
        else:
            return "Medium"
    
    def _suggest_testing_approaches(self, text: str) -> List[str]:
        """Suggest appropriate testing approaches"""
        approaches = []
        
        if any(keyword in text for keyword in ['password', 'login', 'authentication']):
            approaches.extend(['Security Testing', 'Negative Testing', 'Boundary Testing'])
        
        if any(keyword in text for keyword in ['performance', 'speed', 'load']):
            approaches.extend(['Performance Testing', 'Load Testing', 'Stress Testing'])
        
        if any(keyword in text for keyword in ['user interface', 'ui', 'display']):
            approaches.extend(['UI Testing', 'Cross-browser Testing', 'Responsive Testing'])
        
        if any(keyword in text for keyword in ['api', 'integration', 'service']):
            approaches.extend(['Integration Testing', 'API Testing'])
        
        if any(keyword in text for keyword in ['database', 'data', 'storage']):
            approaches.extend(['Data Testing', 'Database Testing'])
        
        # Always include basic approaches
        approaches.extend(['Positive Testing', 'Negative Testing', 'Boundary Testing'])
        
        return list(set(approaches))  # Remove duplicates
    
    def _identify_risk_factors(self, text: str) -> List[str]:
        """Identify potential risk factors"""
        risks = []
        
        if any(keyword in text for keyword in ['security', 'password', 'authentication']):
            risks.append('Security breach risk')
        
        if any(keyword in text for keyword in ['data', 'database', 'storage']):
            risks.append('Data loss/corruption risk')
        
        if any(keyword in text for keyword in ['performance', 'speed', 'load']):
            risks.append('Performance degradation risk')
        
        if any(keyword in text for keyword in ['integration', 'api', 'third party']):
            risks.append('Integration failure risk')
        
        if any(keyword in text for keyword in ['user interface', 'ui', 'display']):
            risks.append('Usability/UX risk')
        
        return risks
    
    def _identify_test_data_needs(self, text: str) -> List[str]:
        """Identify what test data is needed"""
        test_data = []
        
        if 'email' in text:
            test_data.extend(['Valid email addresses', 'Invalid email formats'])
        
        if 'password' in text:
            test_data.extend(['Strong passwords', 'Weak passwords', 'Special characters'])
        
        if 'user' in text:
            test_data.extend(['Valid user accounts', 'Invalid user accounts'])
        
        if any(keyword in text for keyword in ['number', 'numeric', 'integer']):
            test_data.extend(['Valid numbers', 'Invalid numbers', 'Boundary values'])
        
        if any(keyword in text for keyword in ['date', 'time', 'datetime']):
            test_data.extend(['Valid dates', 'Invalid dates', 'Future/past dates'])
        
        return test_data
    
    def _find_integration_points(self, text: str) -> List[str]:
        """Find potential integration points"""
        integrations = []
        
        if any(keyword in text for keyword in ['api', 'service', 'endpoint']):
            integrations.append('External API')
        
        if any(keyword in text for keyword in ['database', 'db', 'storage']):
            integrations.append('Database')
        
        if any(keyword in text for keyword in ['email', 'notification', 'message']):
            integrations.append('Messaging Service')
        
        if any(keyword in text for keyword in ['payment', 'transaction', 'billing']):
            integrations.append('Payment Gateway')
        
        if any(keyword in text for keyword in ['file', 'upload', 'download']):
            integrations.append('File System')
        
        return integrations
    
    def _assess_performance_needs(self, text: str) -> Dict[str, Any]:
        """Assess performance testing needs"""
        needs = {
            'load_testing': False,
            'stress_testing': False,
            'volume_testing': False,
            'response_time_testing': False
        }
        
        if any(keyword in text for keyword in ['concurrent', 'multiple users', 'load']):
            needs['load_testing'] = True
        
        if any(keyword in text for keyword in ['performance', 'speed', 'fast', 'response time']):
            needs['response_time_testing'] = True
        
        if any(keyword in text for keyword in ['large', 'bulk', 'volume', 'massive']):
            needs['volume_testing'] = True
        
        if any(keyword in text for keyword in ['limit', 'maximum', 'capacity']):
            needs['stress_testing'] = True
        
        return needs
    
    def _assess_security_needs(self, text: str) -> Dict[str, Any]:
        """Assess security testing needs"""
        needs = {
            'authentication_testing': False,
            'authorization_testing': False,
            'injection_testing': False,
            'encryption_testing': False,
            'session_testing': False
        }
        
        if any(keyword in text for keyword in ['login', 'authentication', 'password']):
            needs['authentication_testing'] = True
        
        if any(keyword in text for keyword in ['permission', 'access', 'role', 'authorization']):
            needs['authorization_testing'] = True
        
        if any(keyword in text for keyword in ['input', 'form', 'data entry']):
            needs['injection_testing'] = True
        
        if any(keyword in text for keyword in ['encrypt', 'secure', 'protection']):
            needs['encryption_testing'] = True
        
        if any(keyword in text for keyword in ['session', 'token', 'cookie']):
            needs['session_testing'] = True
        
        return needs
    
    def _identify_boundary_conditions(self, text: str) -> List[str]:
        """Identify boundary conditions to test"""
        boundaries = []
        
        # Look for numeric boundaries
        numbers = re.findall(r'\d+', text)
        for num in numbers:
            n = int(num)
            boundaries.extend([
                f"Test with value {n-1} (below limit)",
                f"Test with value {n} (at limit)",
                f"Test with value {n+1} (above limit)"
            ])
        
        # Common boundary scenarios
        if 'length' in text or 'character' in text:
            boundaries.extend(['Minimum length', 'Maximum length', 'Empty string'])
        
        if 'size' in text or 'file' in text:
            boundaries.extend(['Minimum file size', 'Maximum file size', 'Zero size file'])
        
        return boundaries
    
    def _identify_error_scenarios(self, text: str) -> List[str]:
        """Identify potential error scenarios"""
        scenarios = []
        
        if any(keyword in text for keyword in ['login', 'authentication']):
            scenarios.extend([
                'Invalid credentials',
                'Account locked',
                'Session expired',
                'Network timeout'
            ])
        
        if any(keyword in text for keyword in ['file', 'upload']):
            scenarios.extend([
                'File too large',
                'Invalid file format',
                'Corrupted file',
                'No file selected'
            ])
        
        if any(keyword in text for keyword in ['database', 'data']):
            scenarios.extend([
                'Database connection failure',
                'Data validation error',
                'Duplicate data entry',
                'Missing required data'
            ])
        
        return scenarios

class FileProcessor:
    """Handles processing of various file formats containing requirements"""
    
    @staticmethod
    def process_file(file_path: str) -> List[Requirement]:
        """Process a file and extract requirements based on file type"""
        if not os.path.exists(file_path):
            print(f"File not found: {file_path}")
            return []
        
        try:
            file_extension = file_path.lower().split('.')[-1]
            
            if file_extension == 'csv':
                return FileProcessor._process_csv_file(file_path)
            elif file_extension in ['xlsx', 'xls']:
                return FileProcessor._process_excel_file(file_path)
            elif file_extension == 'json':
                return FileProcessor._process_json_file(file_path)
            elif file_extension == 'txt':
                return FileProcessor._process_text_file(file_path)
            elif file_extension in ['docx', 'doc']:
                return FileProcessor._process_word_file(file_path)
            else:
                raise ValueError(f"Unsupported file format: {file_extension}")
                
        except Exception as e:
            print(f"Error processing file {file_path}: {e}")
            return []
    
    @staticmethod
    def _process_csv_file(file_path: str) -> List[Requirement]:
        """Process CSV file containing requirements"""
        requirements = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
        except Exception as e:
            print(f"Error reading CSV file: {e}")
            return []
        
        try:
            # Parse CSV content
            csv_reader = csv.DictReader(io.StringIO(content))
            
            for row_num, row in enumerate(csv_reader, 1):
                # Try different possible column names (case insensitive)
                row_lower = {k.lower(): v for k, v in row.items()}
                
                req_id = (row_lower.get('id') or row_lower.get('requirement id') or 
                         row_lower.get('req_id') or row_lower.get('req id') or 
                         f"REQ{row_num:03d}")
                
                req_text = (row_lower.get('requirement') or row_lower.get('text') or
                           row_lower.get('description') or row_lower.get('details') or "")
                
                priority = (row_lower.get('priority') or "Medium")
                category = (row_lower.get('category') or row_lower.get('type') or "Functional")
                
                if req_text.strip():
                    requirement = Requirement(
                        id=str(req_id),
                        text=req_text.strip(),
                        priority=priority,
                        category=category,
                        source=RequirementSource.CSV_FILE,
                        line_number=row_num
                    )
                    requirements.append(requirement)
        
        except Exception as e:
            print(f"Error parsing CSV content: {e}")
        
        return requirements
    
    @staticmethod
    def _process_excel_file(file_path: str) -> List[Requirement]:
        """Process Excel file containing requirements"""
        requirements = []
        
        try:
            import openpyxl
            
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            worksheet = workbook.active
            
            if worksheet.max_row < 2:
                print("Excel file appears to be empty or has no data rows")
                return []
            
            # Find header row and identify columns
            headers = {}
            for col in range(1, worksheet.max_column + 1):
                header_value = worksheet.cell(1, col).value
                if header_value:
                    header_lower = str(header_value).lower().strip()
                    if any(x in header_lower for x in ['id', 'req']):
                        headers['id'] = col
                    elif any(x in header_lower for x in ['requirement', 'description', 'text', 'detail']):
                        headers['text'] = col
                    elif 'priority' in header_lower:
                        headers['priority'] = col
                    elif any(x in header_lower for x in ['category', 'type']):
                        headers['category'] = col
            
            # Set defaults if headers not found
            if 'id' not in headers:
                headers['id'] = 1
            if 'text' not in headers:
                headers['text'] = 2
            
            # Process data rows
            for row_num in range(2, worksheet.max_row + 1):
                req_id = worksheet.cell(row_num, headers['id']).value
                req_text = worksheet.cell(row_num, headers['text']).value
                priority = worksheet.cell(row_num, headers.get('priority', 3)).value if 'priority' in headers else "Medium"
                category = worksheet.cell(row_num, headers.get('category', 4)).value if 'category' in headers else "Functional"
                
                # Clean up values
                req_id = str(req_id) if req_id else f"REQ{row_num-1:03d}"
                req_text = str(req_text).strip() if req_text else ""
                priority = str(priority) if priority else "Medium"
                category = str(category) if category else "Functional"
                
                if req_text and req_text != "None":
                    requirement = Requirement(
                        id=req_id,
                        text=req_text,
                        priority=priority,
                        category=category,
                        source=RequirementSource.EXCEL_FILE,
                        line_number=row_num
                    )
                    requirements.append(requirement)
                    
        except ImportError:
            print("openpyxl not installed. Install with: pip install openpyxl")
        except Exception as e:
            print(f"Error processing Excel file: {e}")
        
        return requirements
    
    @staticmethod
    def _process_json_file(file_path: str) -> List[Requirement]:
        """Process JSON file containing requirements"""
        requirements = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            
            # Handle different JSON structures
            if isinstance(data, list):
                # Array of requirement objects
                for i, item in enumerate(data):
                    if isinstance(item, dict):
                        req_id = item.get('id', f"REQ{i+1:03d}")
                        req_text = item.get('text', item.get('requirement', item.get('description', "")))
                        priority = item.get('priority', "Medium")
                        category = item.get('category', "Functional")
                        
                        if req_text:
                            requirement = Requirement(
                                id=str(req_id),
                                text=str(req_text),
                                priority=str(priority),
                                category=str(category),
                                source=RequirementSource.JSON_FILE,
                                line_number=i+1
                            )
                            requirements.append(requirement)
            
            elif isinstance(data, dict):
                # Object with requirements key or direct requirement objects
                reqs_data = data.get('requirements', [data] if any(k in data for k in ['id', 'text', 'requirement']) else [])
                
                if isinstance(reqs_data, list):
                    for i, item in enumerate(reqs_data):
                        if isinstance(item, dict):
                            req_id = item.get('id', f"REQ{i+1:03d}")
                            req_text = item.get('text', item.get('requirement', item.get('description', "")))
                            priority = item.get('priority', "Medium")
                            category = item.get('category', "Functional")
                            
                            if req_text:
                                requirement = Requirement(
                                    id=str(req_id),
                                    text=str(req_text),
                                    priority=str(priority),
                                    category=str(category),
                                    source=RequirementSource.JSON_FILE,
                                    line_number=i+1
                                )
                                requirements.append(requirement)
                elif isinstance(reqs_data, dict):
                    # Requirements as key-value pairs
                    for i, (key, value) in enumerate(reqs_data.items()):
                        if isinstance(value, str):
                            req_text = value
                        elif isinstance(value, dict):
                            req_text = value.get('text', value.get('description', str(value)))
                        else:
                            req_text = str(value)
                        
                        requirement = Requirement(
                            id=str(key),
                            text=req_text,
                            priority="Medium",
                            category="Functional",
                            source=RequirementSource.JSON_FILE,
                            line_number=i+1
                        )
                        requirements.append(requirement)
                        
        except Exception as e:
            print(f"Error processing JSON file: {e}")
        
        return requirements
    
    @staticmethod
    def _process_text_file(file_path: str) -> List[Requirement]:
        """Process text file containing requirements"""
        requirements = []
        
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                content = file.read()
            
            lines = content.split('\n')
            req_counter = 1
            current_req = None
            
            for line_num, line in enumerate(lines, 1):
                line = line.strip()
                if not line:
                    continue
                
                # Check if line starts with requirement identifier
                req_match = re.match(r'^(REQ[_-]?\d+|R\d+|\d+\.|\w+\d+):?\s*(.+)', line, re.IGNORECASE)
                if req_match:
                    req_id = req_match.group(1).replace(':', '').strip()
                    req_text = req_match.group(2).strip()
                    
                    if req_text:  # Only create if there's actual text
                        requirement = Requirement(
                            id=req_id,
                            text=req_text,
                            priority="Medium",
                            category="Functional",
                            source=RequirementSource.TEXT_FILE,
                            line_number=line_num
                        )
                        requirements.append(requirement)
                        current_req = requirement
                    
                elif line.startswith(('-', '*', '•', '→')):
                    # Bullet point format
                    req_text = line[1:].strip()
                    if req_text and len(req_text) > 5:  # Minimum length check
                        req_id = f"REQ{req_counter:03d}"
                        req_counter += 1
                        
                        requirement = Requirement(
                            id=req_id,
                            text=req_text,
                            priority="Medium",
                            category="Functional",
                            source=RequirementSource.TEXT_FILE,
                            line_number=line_num
                        )
                        requirements.append(requirement)
                        current_req = requirement
                
                elif re.match(r'^\d+\.\s+', line):
                    # Numbered list format
                    req_text = re.sub(r'^\d+\.\s+', '', line).strip()
                    if req_text and len(req_text) > 5:
                        req_id = f"REQ{req_counter:03d}"
                        req_counter += 1
                        
                        requirement = Requirement(
                            id=req_id,
                            text=req_text,
                            priority="Medium",
                            category="Functional",
                            source=RequirementSource.TEXT_FILE,
                            line_number=line_num
                        )
                        requirements.append(requirement)
                        current_req = requirement
                        
                elif len(line) > 20 and current_req is None:  # Assume longer lines are requirements
                    req_id = f"REQ{req_counter:03d}"
                    req_counter += 1
                    
                    requirement = Requirement(
                        id=req_id,
                        text=line,
                        priority="Medium",
                        category="Functional",
                        source=RequirementSource.TEXT_FILE,
                        line_number=line_num
                    )
                    requirements.append(requirement)
                    current_req = requirement
                    
                elif current_req and len(line) > 10:
                    # Continuation of previous requirement
                    current_req.text += " " + line
                else:
                    current_req = None  # Reset continuation
                    
        except Exception as e:
            print(f"Error processing text file: {e}")
        
        return requirements
    
    @staticmethod
    def _process_word_file(file_path: str) -> List[Requirement]:
        """Process Word document containing requirements"""
        requirements = []
        
        try:
            import docx
            
            doc = docx.Document(file_path)
            req_counter = 1
            
            for para_num, paragraph in enumerate(doc.paragraphs, 1):
                text = paragraph.text.strip()
                if not text or len(text) < 10:
                    continue
                
                # Check if paragraph starts with requirement identifier
                req_match = re.match(r'^(REQ[_-]?\d+|R\d+|\d+\.|\w+\d+):?\s*(.+)', text, re.IGNORECASE)
                if req_match:
                    req_id = req_match.group(1).replace(':', '').strip()
                    req_text = req_match.group(2).strip()
                else:
                    req_id = f"REQ{req_counter:03d}"
                    req_text = text
                    req_counter += 1
                
                # Check if this looks like a requirement
                requirement_indicators = ['shall', 'must', 'will', 'should', 'require', 
                                        'need', 'able to', 'support', 'provide', 'allow',
                                        'enable', 'ensure', 'validate', 'verify']
                
                if any(keyword in text.lower() for keyword in requirement_indicators):
                    requirement = Requirement(
                        id=req_id,
                        text=req_text,
                        priority="Medium",
                        category="Functional",
                        source=RequirementSource.WORD_FILE,
                        line_number=para_num
                    )
                    requirements.append(requirement)
                    
        except ImportError:
            print("python-docx not installed. Install with: pip install python-docx")
        except Exception as e:
            print(f"Error processing Word file: {e}")
        
        return requirements

class TestCaseGenerator:
    """Main class for generating comprehensive test cases from requirements"""
    
    def __init__(self):
        self.analyzer = RequirementAnalyzer()
        self.test_case_counter = 1
    
    def generate_test_cases_from_files(self, file_paths: List[str]) -> Tuple[List[TestCase], Dict[str, Any]]:
        """Generate test cases from uploaded requirement files"""
        all_requirements = []
        file_stats = {}
        
        print(f"📁 Processing {len(file_paths)} requirement files...")
        
        for file_path in file_paths:
            print(f"   📄 Processing {file_path}...")
            requirements = FileProcessor.process_file(file_path)
            
            if requirements:
                print(f"   ✅ Found {len(requirements)} requirements in {file_path}")
                all_requirements.extend(requirements)
                file_stats[file_path] = len(requirements)
            else:
                print(f"   ⚠️ No requirements found in {file_path}")
                file_stats[file_path] = 0
        
        if not all_requirements:
            print("❌ No requirements found in any files")
            return [], file_stats
        
        print(f"📋 Total requirements loaded: {len(all_requirements)}")
        
        # Generate test cases
        test_cases = self._generate_comprehensive_test_cases(all_requirements)
        
        # Generate statistics
        stats = self._generate_statistics(all_requirements, test_cases, file_stats)
        
        return test_cases, stats
    
    def _generate_comprehensive_test_cases(self, requirements: List[Requirement]) -> List[TestCase]:
        """Generate comprehensive test cases from requirements"""
        all_test_cases = []
        
        print("🔍 Analyzing requirements and generating test cases...")
        
        for req in requirements:
            print(f"   Analyzing {req.id}...")
            
            # Analyze the requirement
            analysis = self.analyzer.analyze_requirement(req)
            
            # Generate test cases based on analysis
            test_cases = self._generate_test_cases_for_requirement(req, analysis)
            
            print(f"   ✅ Generated {len(test_cases)} test cases for {req.id}")
            all_test_cases.extend(test_cases)
        
        # Optimize and validate test suite
        optimized_test_cases = self._optimize_test_suite(all_test_cases)
        
        return optimized_test_cases
    
    def _generate_test_cases_for_requirement(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate specific test cases based on requirement analysis"""
        test_cases = []
        req_type = analysis['requirement_type']
        
        # Generate positive test cases
        test_cases.extend(self._generate_positive_test_cases(req, analysis))
        
        # Generate negative test cases
        test_cases.extend(self._generate_negative_test_cases(req, analysis))
        
        # Generate boundary test cases
        test_cases.extend(self._generate_boundary_test_cases(req, analysis))
        
        # Generate type-specific test cases
        if req_type == RequirementType.SECURITY.value:
            test_cases.extend(self._generate_security_test_cases(req, analysis))
        elif req_type == RequirementType.PERFORMANCE.value:
            test_cases.extend(self._generate_performance_test_cases(req, analysis))
        elif req_type == RequirementType.USABILITY.value:
            test_cases.extend(self._generate_usability_test_cases(req, analysis))
        
        # Generate integration test cases if needed
        if analysis['integration_points']:
            test_cases.extend(self._generate_integration_test_cases(req, analysis))
        
        return test_cases
    
    def _generate_positive_test_cases(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate positive test cases (happy path scenarios)"""
        test_cases = []
        
        # Basic positive test case
        test_case = TestCase(
            id=f"{req.id}_POS_{self.test_case_counter:03d}",
            title=f"Verify {req.id} - Valid scenario execution",
            description=f"Test successful execution of requirement: {req.text[:100]}...",
            test_type=TestType.POSITIVE,
            preconditions=self._generate_preconditions(req, analysis, "positive"),
            test_steps=self._generate_test_steps(req, analysis, "positive"),
            expected_result=self._generate_expected_result(req, analysis, "positive"),
            priority=req.priority,
            tags=self._generate_tags(req, analysis, "positive"),
            risk_level=self._assess_risk_level(analysis),
            requirement_id=req.id,
            estimated_time=self._estimate_test_time(analysis, "positive"),
            test_data=analysis.get('test_data_needs', [])
        )
        test_cases.append(test_case)
        self.test_case_counter += 1
        
        # Generate additional positive scenarios based on requirement type
        if 'authentication' in req.text.lower():
            test_cases.extend(self._generate_auth_positive_cases(req, analysis))
        elif 'email' in req.text.lower():
            test_cases.extend(self._generate_email_positive_cases(req, analysis))
        elif 'password' in req.text.lower():
            test_cases.extend(self._generate_password_positive_cases(req, analysis))
        elif 'search' in req.text.lower():
            test_cases.extend(self._generate_search_positive_cases(req, analysis))
        
        return test_cases
    
    def _generate_negative_test_cases(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate negative test cases (error scenarios)"""
        test_cases = []
        
        # Generate test cases for each identified error scenario
        error_scenarios = analysis.get('error_scenarios', [])
        
        for i, error_scenario in enumerate(error_scenarios, 1):
            test_case = TestCase(
                id=f"{req.id}_NEG_{self.test_case_counter:03d}",
                title=f"Verify {req.id} - {error_scenario}",
                description=f"Test system behavior when {error_scenario.lower()}",
                test_type=TestType.NEGATIVE,
                preconditions=self._generate_preconditions(req, analysis, "negative"),
                test_steps=self._generate_error_test_steps(req, error_scenario),
                expected_result=self._generate_error_expected_result(error_scenario),
                priority=req.priority,
                tags=self._generate_tags(req, analysis, "negative"),
                risk_level=self._assess_risk_level(analysis),
                requirement_id=req.id,
                estimated_time=self._estimate_test_time(analysis, "negative"),
                test_data=self._generate_error_test_data(error_scenario)
            )
            test_cases.append(test_case)
            self.test_case_counter += 1
        
        # Generate common negative scenarios
        if not error_scenarios:
            common_negatives = [
                "Invalid input data provided",
                "Required fields left empty",
                "Unauthorized access attempted"
            ]
            
            for scenario in common_negatives:
                test_case = TestCase(
                    id=f"{req.id}_NEG_{self.test_case_counter:03d}",
                    title=f"Verify {req.id} - {scenario}",
                    description=f"Test system behavior when {scenario.lower()}",
                    test_type=TestType.NEGATIVE,
                    preconditions=self._generate_preconditions(req, analysis, "negative"),
                    test_steps=self._generate_error_test_steps(req, scenario),
                    expected_result=self._generate_error_expected_result(scenario),
                    priority="Medium",
                    tags=self._generate_tags(req, analysis, "negative"),
                    risk_level="Medium",
                    requirement_id=req.id,
                    estimated_time=20
                )
                test_cases.append(test_case)
                self.test_case_counter += 1
        
        return test_cases
    
    def _generate_boundary_test_cases(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate boundary test cases"""
        test_cases = []
        
        boundary_conditions = analysis.get('boundary_conditions', [])
        
        for boundary in boundary_conditions:
            test_case = TestCase(
                id=f"{req.id}_BND_{self.test_case_counter:03d}",
                title=f"Verify {req.id} - {boundary}",
                description=f"Test boundary condition: {boundary}",
                test_type=TestType.BOUNDARY,
                preconditions=self._generate_preconditions(req, analysis, "boundary"),
                test_steps=self._generate_boundary_test_steps(req, boundary),
                expected_result=self._generate_boundary_expected_result(boundary),
                priority=req.priority,
                tags=self._generate_tags(req, analysis, "boundary"),
                risk_level=self._assess_risk_level(analysis),
                requirement_id=req.id,
                estimated_time=self._estimate_test_time(analysis, "boundary")
            )
            test_cases.append(test_case)
            self.test_case_counter += 1
        
        return test_cases
    
    def _generate_security_test_cases(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate security-specific test cases"""
        test_cases = []
        
        security_needs = analysis.get('security_considerations', {})
        
        if security_needs.get('authentication_testing'):
            test_case = TestCase(
                id=f"{req.id}_SEC_{self.test_case_counter:03d}",
                title=f"Verify {req.id} - Authentication security",
                description="Test authentication security mechanisms",
                test_type=TestType.SECURITY,
                preconditions=["System is accessible", "Test accounts are available"],
                test_steps=[
                    "Attempt login with valid credentials",
                    "Verify secure session establishment",
                    "Test session timeout behavior",
                    "Verify logout functionality"
                ],
                expected_result="Authentication mechanisms work securely without exposing sensitive data",
                priority="High",
                tags=["security", "authentication"],
                risk_level="High",
                requirement_id=req.id,
                estimated_time=45
            )
            test_cases.append(test_case)
            self.test_case_counter += 1
        
        if security_needs.get('injection_testing'):
            test_case = TestCase(
                id=f"{req.id}_SEC_{self.test_case_counter:03d}",
                title=f"Verify {req.id} - Input validation security",
                description="Test protection against injection attacks",
                test_type=TestType.SECURITY,
                preconditions=["Application is accessible", "Input fields are available"],
                test_steps=[
                    "Enter SQL injection payloads in input fields",
                    "Submit malicious script tags",
                    "Test with special characters and escape sequences",
                    "Verify input sanitization"
                ],
                expected_result="System properly validates and sanitizes all inputs, preventing injection attacks",
                priority="High",
                tags=["security", "injection", "validation"],
                risk_level="High",
                requirement_id=req.id,
                estimated_time=60
            )
            test_cases.append(test_case)
            self.test_case_counter += 1
        
        return test_cases
    
    def _generate_performance_test_cases(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate performance-specific test cases"""
        test_cases = []
        
        perf_needs = analysis.get('performance_considerations', {})
        
        if perf_needs.get('response_time_testing'):
            test_case = TestCase(
                id=f"{req.id}_PERF_{self.test_case_counter:03d}",
                title=f"Verify {req.id} - Response time performance",
                description="Test system response time under normal conditions",
                test_type=TestType.PERFORMANCE,
                preconditions=["System is running", "Performance monitoring tools available"],
                test_steps=[
                    "Execute the functionality with standard input",
                    "Measure response time",
                    "Record performance metrics",
                    "Compare against performance requirements"
                ],
                expected_result="System responds within acceptable time limits",
                priority="High",
                tags=["performance", "response_time"],
                risk_level="Medium",
                requirement_id=req.id,
                estimated_time=30
            )
            test_cases.append(test_case)
            self.test_case_counter += 1
        
        if perf_needs.get('load_testing'):
            test_case = TestCase(
                id=f"{req.id}_PERF_{self.test_case_counter:03d}",
                title=f"Verify {req.id} - Load performance",
                description="Test system behavior under expected load",
                test_type=TestType.PERFORMANCE,
                preconditions=["Load testing environment set up", "Performance baseline established"],
                test_steps=[
                    "Configure load testing parameters",
                    "Execute functionality with multiple concurrent users",
                    "Monitor system resources and response times",
                    "Analyze performance degradation"
                ],
                expected_result="System maintains acceptable performance under expected load",
                priority="High",
                tags=["performance", "load_testing", "scalability"],
                risk_level="Medium",
                requirement_id=req.id,
                estimated_time=90
            )
            test_cases.append(test_case)
            self.test_case_counter += 1
        
        return test_cases
    
    def _generate_usability_test_cases(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate usability-specific test cases"""
        test_cases = []
        
        test_case = TestCase(
            id=f"{req.id}_USA_{self.test_case_counter:03d}",
            title=f"Verify {req.id} - User interface usability",
            description="Test user interface for usability and user experience",
            test_type=TestType.USABILITY,
            preconditions=["Application is accessible", "Different user roles available"],
            test_steps=[
                "Navigate to the relevant interface",
                "Test interface responsiveness",
                "Verify accessibility compliance",
                "Test with different screen resolutions",
                "Evaluate user experience flow"
            ],
            expected_result="Interface is intuitive, accessible, and provides good user experience",
            priority="Medium",
            tags=["usability", "ui", "ux", "accessibility"],
            risk_level="Low",
            requirement_id=req.id,
            estimated_time=45
        )
        test_cases.append(test_case)
        self.test_case_counter += 1
        
        return test_cases
    
    def _generate_integration_test_cases(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate integration-specific test cases"""
        test_cases = []
        
        integration_points = analysis.get('integration_points', [])
        
        for integration in integration_points:
            test_case = TestCase(
                id=f"{req.id}_INT_{self.test_case_counter:03d}",
                title=f"Verify {req.id} - {integration} integration",
                description=f"Test integration with {integration}",
                test_type=TestType.INTEGRATION,
                preconditions=[f"{integration} is available and accessible", "Integration credentials configured"],
                test_steps=[
                    f"Initiate connection to {integration}",
                    "Send test data/request",
                    "Verify successful communication",
                    "Validate response/data exchange",
                    "Test error handling for integration failures"
                ],
                expected_result=f"Successful integration with {integration}, proper data exchange and error handling",
                priority="High",
                tags=["integration", integration.lower().replace(" ", "_")],
                risk_level="High",
                requirement_id=req.id,
                estimated_time=60
            )
            test_cases.append(test_case)
            self.test_case_counter += 1
        
        return test_cases
    
    def _generate_auth_positive_cases(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate positive authentication test cases"""
        test_cases = []
        
        scenarios = [
            ("Valid credentials login", "Test successful login with correct username and password"),
            ("Remember me functionality", "Test remember me option during login"),
            ("Multi-factor authentication", "Test successful MFA process")
        ]
        
        for title_suffix, description in scenarios:
            test_case = TestCase(
                id=f"{req.id}_POS_{self.test_case_counter:03d}",
                title=f"Verify {req.id} - {title_suffix}",
                description=description,
                test_type=TestType.POSITIVE,
                preconditions=["Valid user account exists", "Login page is accessible"],
                test_steps=[
                    "Navigate to login page",
                    "Enter valid credentials",
                    "Complete authentication process",
                    "Verify successful login"
                ],
                expected_result="User successfully authenticates and gains access",
                priority="High",
                tags=["authentication", "positive", "login"],
                risk_level="Medium",
                requirement_id=req.id,
                estimated_time=25
            )
            test_cases.append(test_case)
            self.test_case_counter += 1
        
        return test_cases
    
    def _generate_email_positive_cases(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate positive email validation test cases"""
        test_cases = []
        
        valid_emails = [
            "user@example.com",
            "test.user+tag@domain.co.uk",
            "user123@test-domain.org"
        ]
        
        for email in valid_emails:
            test_case = TestCase(
                id=f"{req.id}_POS_{self.test_case_counter:03d}",
                title=f"Verify {req.id} - Valid email format: {email}",
                description=f"Test email validation with valid format: {email}",
                test_type=TestType.POSITIVE,
                preconditions=["Email input field is available"],
                test_steps=[
                    "Navigate to email input field",
                    f"Enter email: {email}",
                    "Submit or validate the email",
                    "Verify acceptance"
                ],
                expected_result="Email is accepted as valid",
                priority="High",
                tags=["email", "validation", "positive"],
                risk_level="Low",
                requirement_id=req.id,
                estimated_time=15,
                test_data=[email]
            )
            test_cases.append(test_case)
            self.test_case_counter += 1
        
        return test_cases
    
    def _generate_password_positive_cases(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate positive password test cases"""
        test_cases = []
        
        strong_passwords = [
            "StrongPass123!",
            "MySecure@Pass2024",
            "Complex&Password1"
        ]
        
        for password in strong_passwords:
            test_case = TestCase(
                id=f"{req.id}_POS_{self.test_case_counter:03d}",
                title=f"Verify {req.id} - Strong password acceptance",
                description=f"Test password validation with strong password meeting all criteria",
                test_type=TestType.POSITIVE,
                preconditions=["Password input field is available"],
                test_steps=[
                    "Navigate to password input field",
                    "Enter strong password meeting all requirements",
                    "Submit password for validation",
                    "Verify acceptance"
                ],
                expected_result="Strong password is accepted without errors",
                priority="High",
                tags=["password", "validation", "positive", "security"],
                risk_level="Medium",
                requirement_id=req.id,
                estimated_time=20,
                test_data=[password]
            )
            test_cases.append(test_case)
            self.test_case_counter += 1
        
        return test_cases
    
    def _generate_search_positive_cases(self, req: Requirement, analysis: Dict[str, Any]) -> List[TestCase]:
        """Generate positive search test cases"""
        test_cases = []
        
        search_scenarios = [
            ("Exact match search", "Search with exact matching terms"),
            ("Partial match search", "Search with partial matching terms"),
            ("Case insensitive search", "Search with different case variations")
        ]
        
        for scenario, description in search_scenarios:
            test_case = TestCase(
                id=f"{req.id}_POS_{self.test_case_counter:03d}",
                title=f"Verify {req.id} - {scenario}",
                description=description,
                test_type=TestType.POSITIVE,
                preconditions=["Search functionality is available", "Test data exists in system"],
                test_steps=[
                    "Navigate to search interface",
                    "Enter search terms",
                    "Execute search",
                    "Verify results are returned",
                    "Validate result relevance"
                ],
                expected_result="Search returns relevant results matching the search criteria",
                priority="Medium",
                tags=["search", "positive", "functionality"],
                risk_level="Low",
                requirement_id=req.id,
                estimated_time=30
            )
            test_cases.append(test_case)
            self.test_case_counter += 1
        
        return test_cases
    
    def _generate_preconditions(self, req: Requirement, analysis: Dict[str, Any], test_type: str) -> List[str]:
        """Generate appropriate preconditions based on requirement and test type"""
        preconditions = ["System is operational and accessible"]
        
        if 'user' in req.text.lower() or 'login' in req.text.lower():
            preconditions.append("Valid user account is available")
        
        if 'database' in req.text.lower() or 'data' in req.text.lower():
            preconditions.append("Database is accessible and contains test data")
        
        if analysis.get('integration_points'):
            preconditions.append("External systems/integrations are available")
        
        if test_type == "negative":
            preconditions.append("Error handling mechanisms are in place")
        
        if test_type == "performance":
            preconditions.append("Performance monitoring tools are configured")
        
        return preconditions
    
    def _generate_test_steps(self, req: Requirement, analysis: Dict[str, Any], test_type: str) -> List[str]:
        """Generate test steps based on requirement analysis"""
        steps = [
            "Navigate to the relevant functionality",
            "Prepare valid test data",
            "Execute the required operation",
            "Verify system response"
        ]
        
        if 'validation' in req.text.lower():
            steps.insert(-1, "Validate input data format and constraints")
        
        if analysis.get('integration_points'):
            steps.insert(-1, "Verify integration with external systems")
        
        return steps
    
    def _generate_expected_result(self, req: Requirement, analysis: Dict[str, Any], test_type: str) -> str:
        """Generate expected result based on requirement"""
        if test_type == "positive":
            return f"System successfully executes the functionality as specified in {req.id} without errors"
        elif test_type == "negative":
            return "System handles the error condition gracefully with appropriate error messages"
        else:
            return "System behaves according to the specified requirements"
    
    def _generate_error_test_steps(self, req: Requirement, error_scenario: str) -> List[str]:
        """Generate test steps for error scenarios"""
        base_steps = [
            "Navigate to the relevant functionality",
            f"Prepare test data to trigger: {error_scenario}",
            "Execute the operation with invalid/error conditions",
            "Observe system behavior and error handling"
        ]
        return base_steps
    
    def _generate_error_expected_result(self, error_scenario: str) -> str:
        """Generate expected result for error scenarios"""
        return f"System properly handles {error_scenario.lower()} with appropriate error message and graceful degradation"
    
    def _generate_boundary_test_steps(self, req: Requirement, boundary: str) -> List[str]:
        """Generate test steps for boundary conditions"""
        return [
            "Navigate to the relevant functionality",
            f"Prepare test data for boundary condition: {boundary}",
            "Execute the operation with boundary values",
            "Verify system behavior at the boundary",
            "Check for proper validation and handling"
        ]
    
    def _generate_boundary_expected_result(self, boundary: str) -> str:
        """Generate expected result for boundary conditions"""
        return f"System properly handles boundary condition ({boundary}) with correct validation and behavior"
    
    def _generate_tags(self, req: Requirement, analysis: Dict[str, Any], test_type: str) -> List[str]:
        """Generate relevant tags for the test case"""
        tags = [req.id.lower(), test_type, analysis['requirement_type']]
        
        if analysis.get('integration_points'):
            tags.append('integration')
        
        if 'security' in analysis['requirement_type']:
            tags.append('security')
        
        if 'performance' in analysis['requirement_type']:
            tags.append('performance')
        
        return tags
    
    def _assess_risk_level(self, analysis: Dict[str, Any]) -> str:
        """Assess risk level based on requirement analysis"""
        risk_factors = analysis.get('risk_factors', [])
        
        if len(risk_factors) >= 3:
            return "High"
        elif len(risk_factors) >= 1:
            return "Medium"
        else:
            return "Low"
    
    def _estimate_test_time(self, analysis: Dict[str, Any], test_type: str) -> int:
        """Estimate test execution time in minutes"""
        base_time = 30
        
        complexity = analysis.get('complexity', 'Medium')
        if complexity == 'High':
            base_time += 20
        elif complexity == 'Low':
            base_time -= 10
        
        if test_type == 'performance':
            base_time += 30
        elif test_type == 'security':
            base_time += 15
        
        if analysis.get('integration_points'):
            base_time += 20
        
        return max(15, base_time)  # Minimum 15 minutes
    
    def _generate_error_test_data(self, error_scenario: str) -> List[str]:
        """Generate test data for error scenarios"""
        if 'invalid credentials' in error_scenario.lower():
            return ['wrong_username', 'incorrect_password']
        elif 'invalid email' in error_scenario.lower():
            return ['invalid@', '@domain.com', 'not-an-email']
        elif 'file too large' in error_scenario.lower():
            return ['large_file.zip (>10MB)']
        else:
            return ['invalid_input', 'malformed_data']
    
    def _optimize_test_suite(self, test_cases: List[TestCase]) -> List[TestCase]:
        """Optimize the test suite by removing duplicates and prioritizing"""
        print("⚡ Optimizing test suite...")
        
        # Remove duplicates based on title similarity
        unique_cases = []
        seen_titles = set()
        
        for case in test_cases:
            title_key = case.title.lower().replace(" ", "")
            if title_key not in seen_titles:
                seen_titles.add(title_key)
                unique_cases.append(case)
        
        # Prioritize test cases
        priority_order = {"High": 1, "Medium": 2, "Low": 3}
        type_order = {
            TestType.POSITIVE: 1,
            TestType.SECURITY: 2,
            TestType.NEGATIVE: 3,
            TestType.BOUNDARY: 4,
            TestType.PERFORMANCE: 5,
            TestType.INTEGRATION: 6,
            TestType.USABILITY: 7,
            TestType.EDGE_CASE: 8
        }
        
        optimized_cases = sorted(unique_cases, key=lambda x: (
            priority_order.get(x.priority, 2),
            type_order.get(x.test_type, 5)
        ))
        
        print(f"   Optimized from {len(test_cases)} to {len(optimized_cases)} test cases")
        
        return optimized_cases
    
    def _generate_statistics(self, requirements: List[Requirement], test_cases: List[TestCase], file_stats: Dict[str, int]) -> Dict[str, Any]:
        """Generate comprehensive statistics"""
        stats = {
            'summary': {
                'total_requirements': len(requirements),
                'total_test_cases': len(test_cases),
                'files_processed': len(file_stats),
                'avg_test_cases_per_requirement': round(len(test_cases) / len(requirements), 2) if requirements else 0
            },
            'by_file': file_stats,
            'by_requirement_type': {},
            'by_test_type': {},
            'by_priority': {},
            'by_risk_level': {},
            'estimated_execution_time': {
                'total_minutes': sum(tc.estimated_time for tc in test_cases),
                'total_hours': round(sum(tc.estimated_time for tc in test_cases) / 60, 2)
            },
            'coverage_analysis': {
                'requirements_with_security_tests': 0,
                'requirements_with_performance_tests': 0,
                'requirements_with_integration_tests': 0
            }
        }
        
        # Analyze by requirement type
        for req in requirements:
            analyzer = RequirementAnalyzer()
            req_type = analyzer._determine_requirement_type(req.text.lower())
            stats['by_requirement_type'][req_type] = stats['by_requirement_type'].get(req_type, 0) + 1
        
        # Analyze by test type
        for tc in test_cases:
            test_type = tc.test_type.value
            stats['by_test_type'][test_type] = stats['by_test_type'].get(test_type, 0) + 1
        
        # Analyze by priority
        for tc in test_cases:
            priority = tc.priority
            stats['by_priority'][priority] = stats['by_priority'].get(priority, 0) + 1
        
        # Analyze by risk level
        for tc in test_cases:
            risk = tc.risk_level
            stats['by_risk_level'][risk] = stats['by_risk_level'].get(risk, 0) + 1
        
        # Coverage analysis
        req_ids_with_security = set()
        req_ids_with_performance = set()
        req_ids_with_integration = set()
        
        for tc in test_cases:
            if tc.test_type == TestType.SECURITY:
                req_ids_with_security.add(tc.requirement_id)
            elif tc.test_type == TestType.PERFORMANCE:
                req_ids_with_performance.add(tc.requirement_id)
            elif tc.test_type == TestType.INTEGRATION:
                req_ids_with_integration.add(tc.requirement_id)
        
        stats['coverage_analysis']['requirements_with_security_tests'] = len(req_ids_with_security)
        stats['coverage_analysis']['requirements_with_performance_tests'] = len(req_ids_with_performance)
        stats['coverage_analysis']['requirements_with_integration_tests'] = len(req_ids_with_integration)
        
        return stats
    
    def export_to_json(self, test_cases: List[TestCase], filename: str = "test_cases.json"):
        """Export test cases to JSON"""
        cases_dict = []
        for case in test_cases:
            case_dict = asdict(case)
            case_dict["test_type"] = case.test_type.value
            cases_dict.append(case_dict)
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(cases_dict, f, indent=2, ensure_ascii=False)
        
        print(f"✅ Exported {len(test_cases)} test cases to {filename}")
    
    def export_to_csv(self, test_cases: List[TestCase], filename: str = "test_cases.csv"):
        """Export test cases to CSV format"""
        with open(filename, 'w', newline='', encoding='utf-8') as csvfile:
            fieldnames = ['Test ID', 'Title', 'Description', 'Type', 'Priority', 
                         'Preconditions', 'Test Steps', 'Expected Result', 'Tags',
                         'Risk Level', 'Requirement ID', 'Automation Feasible',
                         'Estimated Time (min)', 'Test Data']
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            
            writer.writeheader()
            for tc in test_cases:
                writer.writerow({
                    'Test ID': tc.id,
                    'Title': tc.title,
                    'Description': tc.description,
                    'Type': tc.test_type.value,
                    'Priority': tc.priority,
                    'Preconditions': '; '.join(tc.preconditions),
                    'Test Steps': '; '.join(tc.test_steps),
                    'Expected Result': tc.expected_result,
                    'Tags': ', '.join(tc.tags),
                    'Risk Level': tc.risk_level,
                    'Requirement ID': tc.requirement_id,
                    'Automation Feasible': tc.automation_feasible,
                    'Estimated Time (min)': tc.estimated_time,
                    'Test Data': '; '.join(tc.test_data)
                })
        
        print(f"✅ Exported {len(test_cases)} test cases to {filename}")
    
    def export_to_excel(self, test_cases: List[TestCase], filename: str = "test_cases.xlsx"):
        """Export test cases to Excel format"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
            
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Test Cases"
            
            # Headers
            headers = ['Test ID', 'Title', 'Description', 'Type', 'Priority', 
                      'Preconditions', 'Test Steps', 'Expected Result', 'Tags',
                      'Risk Level', 'Requirement ID', 'Automation Feasible',
                      'Estimated Time (min)', 'Test Data']
            
            # Style definitions
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            border = Border(left=Side(style='thin'), right=Side(style='thin'),
                           top=Side(style='thin'), bottom=Side(style='thin'))
            
            # Add headers with styling
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = border
            
            # Add data
            for row, tc in enumerate(test_cases, 2):
                data_row = [
                    tc.id,
                    tc.title,
                    tc.description,
                    tc.test_type.value,
                    tc.priority,
                    '; '.join(tc.preconditions),
                    '; '.join(tc.test_steps),
                    tc.expected_result,
                    ', '.join(tc.tags),
                    tc.risk_level,
                    tc.requirement_id,
                    tc.automation_feasible,
                    tc.estimated_time,
                    '; '.join(tc.test_data)
                ]
                
                for col, value in enumerate(data_row, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                    
                    # Color code by priority
                    if tc.priority == "High":
                        cell.fill = PatternFill(start_color="FFE6E6", end_color="FFE6E6", fill_type="solid")
                    elif tc.priority == "Medium":
                        cell.fill = PatternFill(start_color="FFF2E6", end_color="FFF2E6", fill_type="solid")
                    else:  # Low
                        cell.fill = PatternFill(start_color="E6F3E6", end_color="E6F3E6", fill_type="solid")
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)  # Cap at 50 for readability
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Set row heights for better readability
            for row in range(2, len(test_cases) + 2):
                ws.row_dimensions[row].height = 40
            
            # Add summary sheet
            ws_summary = wb.create_sheet(title="Summary")
            
            # Summary data
            summary_data = [
                ["Metric", "Value"],
                ["Total Test Cases", len(test_cases)],
                ["High Priority", sum(1 for tc in test_cases if tc.priority == "High")],
                ["Medium Priority", sum(1 for tc in test_cases if tc.priority == "Medium")],
                ["Low Priority", sum(1 for tc in test_cases if tc.priority == "Low")],
                ["", ""],
                ["Test Types", "Count"],
                ["Positive Tests", sum(1 for tc in test_cases if tc.test_type == TestType.POSITIVE)],
                ["Negative Tests", sum(1 for tc in test_cases if tc.test_type == TestType.NEGATIVE)],
                ["Boundary Tests", sum(1 for tc in test_cases if tc.test_type == TestType.BOUNDARY)],
                ["Security Tests", sum(1 for tc in test_cases if tc.test_type == TestType.SECURITY)],
                ["Performance Tests", sum(1 for tc in test_cases if tc.test_type == TestType.PERFORMANCE)],
                ["Integration Tests", sum(1 for tc in test_cases if tc.test_type == TestType.INTEGRATION)],
                ["", ""],
                ["Estimated Execution Time", ""],
                ["Total Minutes", sum(tc.estimated_time for tc in test_cases)],
                ["Total Hours", round(sum(tc.estimated_time for tc in test_cases) / 60, 2)]
            ]
            
            for row, (metric, value) in enumerate(summary_data, 1):
                ws_summary.cell(row=row, column=1, value=metric).font = Font(bold=True)
                ws_summary.cell(row=row, column=2, value=value)
            
            # Auto-adjust summary sheet columns
            for column in ws_summary.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                ws_summary.column_dimensions[column_letter].width = max_length + 2
            
            # Save workbook
            wb.save(filename)
            print(f"✅ Exported {len(test_cases)} test cases to {filename}")
            
        except ImportError:
            print("⚠️ openpyxl not installed. Install with: pip install openpyxl")
            # Fallback to CSV
            csv_filename = filename.replace('.xlsx', '.csv')
            self.export_to_csv(test_cases, csv_filename)
        except Exception as e:
            print(f"❌ Error exporting to Excel: {e}")
            # Fallback to CSV
            csv_filename = filename.replace('.xlsx', '.csv')
            self.export_to_csv(test_cases, csv_filename)
    
    def generate_test_report(self, test_cases: List[TestCase], stats: Dict[str, Any], filename: str = "test_report.html"):
        """Generate an HTML test report"""
        html_content = f"""
        <!DOCTYPE html>
        <html lang="en">
        <head>
            <meta charset="UTF-8">
            <meta name="viewport" content="width=device-width, initial-scale=1.0">
            <title>Test Case Generation Report</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; background-color: #f5f5f5; }}
                .container {{ max-width: 1200px; margin: 0 auto; background-color: white; padding: 20px; border-radius: 8px; box-shadow: 0 2px 4px rgba(0,0,0,0.1); }}
                h1, h2, h3 {{ color: #333; }}
                .summary {{ background-color: #e8f4f8; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
                .stats {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(250px, 1fr)); gap: 15px; margin-bottom: 20px; }}
                .stat-card {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; border-left: 4px solid #007bff; }}
                .test-case {{ background-color: #ffffff; border: 1px solid #ddd; border-radius: 5px; padding: 15px; margin-bottom: 15px; }}
                .test-case h4 {{ margin-top: 0; color: #495057; }}
                .priority-high {{ border-left: 4px solid #dc3545; }}
                .priority-medium {{ border-left: 4px solid #ffc107; }}
                .priority-low {{ border-left: 4px solid #28a745; }}
                .test-type {{ display: inline-block; padding: 3px 8px; border-radius: 3px; font-size: 12px; color: white; margin-right: 5px; }}
                .type-positive {{ background-color: #28a745; }}
                .type-negative {{ background-color: #dc3545; }}
                .type-boundary {{ background-color: #17a2b8; }}
                .type-security {{ background-color: #6f42c1; }}
                .type-performance {{ background-color: #fd7e14; }}
                .type-integration {{ background-color: #20c997; }}
                .type-usability {{ background-color: #e83e8c; }}
                table {{ width: 100%; border-collapse: collapse; margin-top: 10px; }}
                th, td {{ border: 1px solid #ddd; padding: 8px; text-align: left; }}
                th {{ background-color: #f2f2f2; }}
            </style>
        </head>
        <body>
            <div class="container">
                <h1>Test Case Generation Report</h1>
                <p>Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
                
                <div class="summary">
                    <h2>Summary</h2>
                    <div class="stats">
                        <div class="stat-card">
                            <h3>{stats['summary']['total_requirements']}</h3>
                            <p>Total Requirements</p>
                        </div>
                        <div class="stat-card">
                            <h3>{stats['summary']['total_test_cases']}</h3>
                            <p>Total Test Cases</p>
                        </div>
                        <div class="stat-card">
                            <h3>{stats['summary']['files_processed']}</h3>
                            <p>Files Processed</p>
                        </div>
                        <div class="stat-card">
                            <h3>{stats['estimated_execution_time']['total_hours']} hrs</h3>
                            <p>Estimated Execution Time</p>
                        </div>
                    </div>
                </div>
                
                <h2>Statistics</h2>
                
                <h3>Test Cases by Type</h3>
                <table>
                    <tr><th>Type</th><th>Count</th><th>Percentage</th></tr>"""
        
        for test_type, count in stats['by_test_type'].items():
            percentage = round((count / stats['summary']['total_test_cases']) * 100, 1)
            html_content += f"<tr><td>{test_type.title()}</td><td>{count}</td><td>{percentage}%</td></tr>"
        
        html_content += """
                </table>
                
                <h3>Test Cases by Priority</h3>
                <table>
                    <tr><th>Priority</th><th>Count</th><th>Percentage</th></tr>"""
        
        for priority, count in stats['by_priority'].items():
            percentage = round((count / stats['summary']['total_test_cases']) * 100, 1)
            html_content += f"<tr><td>{priority}</td><td>{count}</td><td>{percentage}%</td></tr>"
        
        html_content += """
                </table>
                
                <h2>Test Cases</h2>"""
        
        # Group test cases by requirement
        test_cases_by_req = {}
        for tc in test_cases:
            req_id = tc.requirement_id
            if req_id not in test_cases_by_req:
                test_cases_by_req[req_id] = []
            test_cases_by_req[req_id].append(tc)
        
        for req_id, cases in test_cases_by_req.items():
            html_content += f"<h3>Requirement: {req_id}</h3>"
            
            for tc in cases:
                priority_class = f"priority-{tc.priority.lower()}"
                type_class = f"type-{tc.test_type.value}"
                
                html_content += f"""
                <div class="test-case {priority_class}">
                    <h4>{tc.title}</h4>
                    <span class="test-type {type_class}">{tc.test_type.value.upper()}</span>
                    <span class="test-type" style="background-color: #6c757d;">Priority: {tc.priority}</span>
                    <span class="test-type" style="background-color: #6c757d;">Risk: {tc.risk_level}</span>
                    <span class="test-type" style="background-color: #6c757d;">Time: {tc.estimated_time}min</span>
                    
                    <p><strong>Description:</strong> {tc.description}</p>
                    
                    <p><strong>Preconditions:</strong></p>
                    <ul>"""
                
                for precondition in tc.preconditions:
                    html_content += f"<li>{precondition}</li>"
                
                html_content += """</ul>
                    
                    <p><strong>Test Steps:</strong></p>
                    <ol>"""
                
                for step in tc.test_steps:
                    html_content += f"<li>{step}</li>"
                
                html_content += f"""</ol>
                    
                    <p><strong>Expected Result:</strong> {tc.expected_result}</p>
                    
                    <p><strong>Tags:</strong> {', '.join(tc.tags)}</p>
                </div>"""
        
        html_content += """
            </div>
        </body>
        </html>"""
        
        with open(filename, 'w', encoding='utf-8') as f:
            f.write(html_content)
        
        print(f"✅ Generated HTML report: {filename}")


def main():
    """Example usage of the Test Case Generator with real file processing"""
    
    def create_sample_files():
        """Create sample requirement files for demonstration"""
        
        # Sample CSV file
        csv_content = """ID,Requirement,Priority,Category
REQ001,"The system shall validate user email addresses using RFC 5322 standard and reject invalid formats with appropriate error messages",High,Authentication
REQ002,"User passwords must be between 8 and 128 characters long and contain at least one uppercase letter, one lowercase letter, one digit, and one special character",High,Security
REQ003,"The system shall implement rate limiting to prevent brute force attacks, allowing maximum 5 login attempts per IP address within 15 minutes",High,Security
REQ004,"When a user enters invalid credentials 3 consecutive times, the account shall be temporarily locked for 30 minutes",High,Security
REQ005,"The system shall log all authentication events including successful logins, failed attempts, and account lockouts for security audit purposes",Medium,Logging
REQ006,"The search functionality shall return results within 2 seconds for queries with less than 1000 matching records",High,Performance
REQ007,"The user interface shall be responsive and accessible on mobile devices with screen sizes from 320px to 768px width",Medium,Usability
REQ008,"The system shall integrate with external payment gateway API and handle transaction responses within 30 seconds",High,Integration
REQ009,"User profile data shall be encrypted using AES-256 encryption when stored in the database",High,Security
REQ010,"The application shall support concurrent access for up to 10,000 simultaneous users without performance degradation",High,Performance"""
        
        with open('sample_requirements.csv', 'w', encoding='utf-8') as f:
            f.write(csv_content)
        
        # Sample JSON file
        json_content = {
            "requirements": [
                {
                    "id": "REQ011",
                    "text": "The system shall provide a password reset functionality via email verification with secure token expiration within 24 hours",
                    "priority": "Medium",
                    "category": "Authentication"
                },
                {
                    "id": "REQ012", 
                    "text": "Users shall be able to enable two-factor authentication using TOTP (Time-based One-Time Password) with QR code generation",
                    "priority": "High",
                    "category": "Security"
                },
                {
                    "id": "REQ013",
                    "text": "The system shall automatically backup user data daily at 2 AM with verification of backup integrity",
                    "priority": "High",
                    "category": "Data Management"
                },
                {
                    "id": "REQ014",
                    "text": "API endpoints shall return standardized error responses with appropriate HTTP status codes and detailed error messages",
                    "priority": "Medium",
                    "category": "Integration"
                }
            ]
        }
        
        with open('sample_requirements.json', 'w', encoding='utf-8') as f:
            json.dump(json_content, f, indent=2)
        
        # Sample text file with various formats
        text_content = """Requirements Document

REQ015: The system shall maintain user session for maximum 8 hours of inactivity before automatic logout
REQ016: All user inputs shall be sanitized to prevent SQL injection attacks and XSS vulnerabilities

Security Requirements:
- The system must encrypt all sensitive data at rest using AES-256 encryption
- User profile data shall be editable by authenticated users only with proper authorization checks
- API access requires valid JWT tokens with expiration time not exceeding 24 hours

Performance Requirements:
REQ020: The application shall support concurrent users up to 10,000 simultaneous sessions
REQ021: Database queries shall be optimized to execute within 500ms for 95% of requests

1. The file upload feature shall accept files up to 50MB in size
2. Uploaded files shall be scanned for malware before processing
3. The system shall generate thumbnails for image files within 10 seconds of upload

Integration Requirements:
REQ025: The system shall integrate with LDAP for user authentication and authorization
REQ026: Email notifications shall be sent via SMTP server with retry mechanism for failed deliveries"""
        
        with open('sample_requirements.txt', 'w', encoding='utf-8') as f:
            f.write(text_content)
        
        return ['sample_requirements.csv', 'sample_requirements.json', 'sample_requirements.txt']
    
    print("🤖 Advanced Test Case Generator with Real File Processing")
    print("=" * 60)
    
    generator = TestCaseGenerator()
    
    print("\n🔧 Creating comprehensive sample requirement files...")
    sample_files = create_sample_files()
    print(f"   ✅ Created: {', '.join(sample_files)}")
    
    try:
        print(f"\n📂 Processing requirement files...")
        # Generate test cases from files
        test_cases, stats = generator.generate_test_cases_from_files(sample_files)
        
        if not test_cases:
            print("❌ No test cases generated. Please check your requirement files.")
            return
        
        # Display results
        print(f"\n🎉 Successfully generated {len(test_cases)} test cases!")
        
        # Print detailed statistics
        print(f"\n📊 Detailed Statistics:")
        print(f"   📋 Requirements processed: {stats['summary']['total_requirements']}")
        print(f"   📝 Test cases generated: {stats['summary']['total_test_cases']}")
        print(f"   📁 Files processed: {stats['summary']['files_processed']}")
        print(f"   ⭐ Avg test cases per requirement: {stats['summary']['avg_test_cases_per_requirement']}")
        print(f"   ⏱️ Total estimated execution time: {stats['estimated_execution_time']['total_hours']} hours")
        
        print(f"\n📈 Test Cases by Type:")
        for test_type, count in stats['by_test_type'].items():
            percentage = round((count / stats['summary']['total_test_cases']) * 100, 1)
            print(f"   {test_type.title()}: {count} ({percentage}%)")
        
        print(f"\n🎯 Test Cases by Priority:")
        for priority, count in stats['by_priority'].items():
            percentage = round((count / stats['summary']['total_test_cases']) * 100, 1)
            print(f"   {priority}: {count} ({percentage}%)")
        
        print(f"\n⚠️ Test Cases by Risk Level:")
        for risk, count in stats['by_risk_level'].items():
            percentage = round((count / stats['summary']['total_test_cases']) * 100, 1)
            print(f"   {risk}: {count} ({percentage}%)")
        
        # Show sample test cases for different types
        print(f"\n🔐 Sample Security Test Case:")
        security_cases = [tc for tc in test_cases if tc.test_type == TestType.SECURITY]
        if security_cases:
            sample_case = security_cases[0]
            print(f"   ID: {sample_case.id}")
            print(f"   Title: {sample_case.title}")
            print(f"   Priority: {sample_case.priority}, Risk: {sample_case.risk_level}")
            print(f"   Estimated Time: {sample_case.estimated_time} minutes")
        
        print(f"\n⚡ Sample Performance Test Case:")
        perf_cases = [tc for tc in test_cases if tc.test_type == TestType.PERFORMANCE]
        if perf_cases:
            sample_case = perf_cases[0]
            print(f"   ID: {sample_case.id}")
            print(f"   Title: {sample_case.title}")
            print(f"   Priority: {sample_case.priority}, Risk: {sample_case.risk_level}")
            print(f"   Estimated Time: {sample_case.estimated_time} minutes")
        
        # Export results in different formats
        print(f"\n💾 Exporting test cases to multiple formats...")
        
        # Export to JSON
        generator.export_to_json(test_cases, "comprehensive_test_cases.json")
        
        # Export to CSV
        generator.export_to_csv(test_cases, "comprehensive_test_cases.csv")
        
        # Export to Excel (if openpyxl is available)
        generator.export_to_excel(test_cases, "comprehensive_test_cases.xlsx")
        
        # Generate HTML report
        generator.generate_test_report(test_cases, stats, "test_case_report.html")
        
        print(f"\n📚 Generated Files:")
        print(f"   📄 JSON: comprehensive_test_cases.json")
        print(f"   📄 CSV: comprehensive_test_cases.csv") 
        print(f"   📄 Excel: comprehensive_test_cases.xlsx")
        print(f"   📄 HTML Report: test_case_report.html")
        
        # Show coverage analysis
        print(f"\n🎯 Coverage Analysis:")
        coverage = stats['coverage_analysis']
        total_reqs = stats['summary']['total_requirements']
        print(f"   🔒 Requirements with Security Tests: {coverage['requirements_with_security_tests']}/{total_reqs}")
        print(f"   ⚡ Requirements with Performance Tests: {coverage['requirements_with_performance_tests']}/{total_reqs}")
        print(f"   🔗 Requirements with Integration Tests: {coverage['requirements_with_integration_tests']}/{total_reqs}")
        
        # Usage examples
        print(f"\n💡 Usage Examples:")
        print(f"   📁 Single file: generator.generate_test_cases_from_files(['requirements.csv'])")
        print(f"   📁 Multiple files: generator.generate_test_cases_from_files(['req1.csv', 'req2.xlsx', 'req3.json'])")
        print(f"   📄 Mixed formats: Supports CSV, Excel, JSON, TXT, and Word documents")
        
        # Clean up sample files
        print(f"\n🧹 Cleaning up sample files...")
        for file in sample_files:
            try:
                os.remove(file)
                print(f"   ✅ Removed: {file}")
            except:
                print(f"   ⚠️ Could not remove: {file}")
        
        print(f"\n✨ Test case generation completed successfully!")
        print(f"📖 Open 'test_case_report.html' in your browser to view the detailed report.")
        
    except Exception as e:
        print(f"❌ Error during test case generation: {e}")
        import traceback
        traceback.print_exc()

if __name__ == "__main__":
    main()